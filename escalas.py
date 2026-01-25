# -*- coding: utf-8 -*-
"""
escalas.py (sin pandas)
Lee el maestro Excel con openpyxl y expone helpers para:
- /meta (ramas, meses, agrupamientos, categorias)
- /payload (básico / no_rem / suma_fija)
- reglas: conexiones Agua Potable, adicionales Fúnebres, título Turismo, cajero, KM
"""
from __future__ import annotations

import os
import re
import datetime as _dt
from functools import lru_cache
from typing import Dict, Tuple, List, Any, Optional

from decimal import Decimal, ROUND_HALF_UP

import openpyxl


# ---------------------------
# Config
# ---------------------------

def _default_maestro_path() -> str:
    # Preferimos SIEMPRE data/maestro_actualizado.xlsx (evita conflicto con un maestro en raíz).
    # Resolvemos relativo a este archivo, no al CWD, para evitar errores 500 en deploy.
    here = os.path.dirname(__file__)

    p1 = os.path.join(here, "data", "maestro_actualizado.xlsx")
    if os.path.exists(p1):
        return p1

    p2 = os.path.join(here, "data", "maestro.xlsx")
    if os.path.exists(p2):
        return p2

    # último fallback (raíz del proyecto)
    p3 = os.path.join(here, "maestro.xlsx")
    if os.path.exists(p3):
        return p3

    return "maestro.xlsx"


MAESTRO_PATH = os.getenv("MAESTRO_PATH", _default_maestro_path())


def round2(x: float) -> float:
    """Redondeo a 2 decimales (half up) para importes."""
    try:
        return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0.0


# ---------------------------
# Utils
# ---------------------------

def _mes_to_key(v: Any) -> str:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m")
    if v is None:
        return ""
    s = str(v).strip()
    # admite "2026-04-01 00:00:00"
    if len(s) >= 7 and s[4] == "-" and s[6].isdigit():
        return s[:7]
    return s

def _to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # números argentinos: "1.176.516" o "1.176.516,50"
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def _norm(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def norm_rama(rama: Any) -> str:
    """Normaliza el nombre de rama para comparaciones."""
    return _norm(rama).upper().replace("  ", " ").strip()



def _extract_hs_from_categoria(cat: Any) -> Optional[float]:
    """Extrae horas de una categoría tipo '... 20HS' / '... 36 hs'."""
    s = _norm(cat).upper()
    if not s:
        return None
    m = re.search(r"(\d+(?:[\.,]\d+)?)\s*H", s)
    if not m:
        return None
    raw = m.group(1).replace(",", ".")
    try:
        v = float(raw)
    except Exception:
        return None
    # límites razonables
    if v <= 0 or v > 48:
        return None
    return v

def _nr_labels(rama: str) -> dict:
    """Nombres oficiales de los NR según rama (criterio César)."""
    r = _norm(rama).upper()
    if r in ("TURISMO", "CEREALES"):
        # En el maestro de Turismo/Cereales, suele venir 60k en no_rem y 40k en suma_fija.
        return {
            "no_rem": "Recomp. NR. Acu. 26",
            "suma_fija": "Incr. NR. Acu. Ene 26",
        }
    return {
        "no_rem": "Incr. NR. Acu. Dic 25",
        "suma_fija": "Recomp. NR. Acu. 25",
    }

# ---------------------------
# Maestro loader / parser
# ---------------------------

@lru_cache(maxsize=1)
def _load_wb() -> openpyxl.Workbook:
    return openpyxl.load_workbook(MAESTRO_PATH, data_only=True)

@lru_cache(maxsize=1)
def _build_index() -> Dict[str, Any]:
    wb = _load_wb()

    # salida
    payload: Dict[Tuple[str, str, str, str], Dict[str, float]] = {}
    ramas_set = set()
    meses_set = set()
    agrup_by_rama: Dict[str, set] = {}
    cat_by_rama_agrup: Dict[Tuple[str, str], set] = {}
    meses_by_rama: Dict[str, set] = {}

    def add_row(rama: str, agrup: str, cat: str, mes: str, bas: float, nr: float, sf: float):
        rama_u = _norm(rama).upper()
        agrup_u = _norm(agrup) if _norm(agrup) else "—"
        cat_u = _norm(cat) if _norm(cat) else "—"

        # Fix maestro FUNEBRES: a veces las categorías quedaron en "Agrupamiento" y "Categoria" viene vacío.
        if rama_u in ("FUNEBRES", "FÚNEBRES") and (cat_u == "—" or cat_u == "") and agrup_u not in ("—", ""):
            cat_u = agrup_u
            agrup_u = "—"
        mes_k = _mes_to_key(mes)

        if not rama_u or not mes_k:
            return

        payload[(rama_u, agrup_u, cat_u, mes_k)] = {"basico": bas, "no_rem": nr, "suma_fija": sf}
        ramas_set.add(rama_u)
        meses_set.add(mes_k)
        agrup_by_rama.setdefault(rama_u, set()).add(agrup_u)
        cat_by_rama_agrup.setdefault((rama_u, agrup_u), set()).add(cat_u)
        meses_by_rama.setdefault(rama_u, set()).add(mes_k)

    # --- Tabulares (GENERAL, TURISMO, FUNEBRES, CEREALES, CALL CENTER)
    for sh_name in wb.sheetnames:
        if not sh_name.startswith("Categorias_"):
            continue
        if sh_name == "Categorias_Agua_Potable":
            continue  # parse especial abajo

        ws = wb[sh_name]
        # headers en fila 1
        headers = [_norm(ws.cell(1, c).value).lower() for c in range(1, 10)]
        # buscamos indices
        def idx(name: str) -> Optional[int]:
            for i,h in enumerate(headers, start=1):
                if h == name:
                    return i
            return None

        i_rama = idx("rama") or 1
        i_agr = idx("agrupamiento") or 2
        i_cat = idx("categoria") or 3
        i_mes = idx("mes") or 4
        i_bas = idx("basico") or 5
        i_nr  = idx("no_rem") or 6
        i_sf  = idx("suma_fija") or 7

        for r in range(2, ws.max_row + 1):
            rama = ws.cell(r, i_rama).value
            if rama is None:
                continue
            mes = ws.cell(r, i_mes).value
            rama_u = _norm(rama).upper()
            agrup = ws.cell(r, i_agr).value
            cat = ws.cell(r, i_cat).value
            bas = _to_float(ws.cell(r, i_bas).value)
            nr  = _to_float(ws.cell(r, i_nr).value)
            sf  = _to_float(ws.cell(r, i_sf).value)
            add_row(rama_u, agrup, cat, mes, bas, nr, sf)

    # --- AGUA POTABLE (sheet no tabular, por bloques)
    if "Categorias_Agua_Potable" in wb.sheetnames:
        ws = wb["Categorias_Agua_Potable"]
        rama_u = "AGUA POTABLE"
        current_agr = "—"
        current_cat = ""
        in_table = False

        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value

            a_s = _norm(a)

            # AGRUPAMIENTO:
            if isinstance(a, str) and a_s.upper().startswith("AGRUPAMIENTO"):
                # el valor puede venir en col 2
                current_agr = _norm(b) if _norm(b) else "—"
                in_table = False
                continue

            # Categoría:
            if isinstance(a, str) and a_s.upper().startswith("CATEGOR"):
                current_cat = _norm(b)
                in_table = False
                continue

            # header MES - AÑO
            if isinstance(a, str) and a_s.upper().startswith("MES"):
                in_table = True
                continue

            if not in_table:
                continue

            # filas de mes
            mes_k = _mes_to_key(a)
            if not mes_k or mes_k.lower().startswith("mes"):
                continue

            bas = _to_float(b)
            # En Agua Potable, los NR vienen como 2 columnas (incrementos NR) => los consolidamos en "suma_fija"
            sf = _to_float(c) + _to_float(d)
            nr = 0.0

            add_row(rama_u, current_agr or "—", current_cat or "—", mes_k, bas, nr, sf)

    # ---------------------------
    # Adicionales Fúnebres
    # ---------------------------
    # La hoja "Adicionales" del maestro puede venir en distintos formatos según versión.
    # Formato usual actual (enero 2026): Rama | Concepto | Mes | Valor | Detalle
    # Otros formatos posibles: Rama | Concepto | Mes | Tipo | Monto | % | Observación
    funebres_adic: Dict[str, List[Dict[str, Any]]] = {}  # mes -> list
    if "Adicionales" in wb.sheetnames:
        ws = wb["Adicionales"]

        # Mapear columnas por encabezados (fila 1)
        header = {}
        for c in range(1, ws.max_column + 1):
            h = _norm(ws.cell(1, c).value)
            if h:
                header[h.lower()] = c

        col_rama = header.get("rama", 1)
        col_concepto = header.get("concepto", 2)
        col_mes = header.get("mes", 3)
        col_tipo = header.get("tipo")  # opcional
        col_monto = header.get("monto") or header.get("valor") or header.get("importe")
        col_pct = header.get("%") or header.get("porcentaje") or header.get("pct")
        col_obs = header.get("observación") or header.get("observacion") or header.get("detalle") or header.get("obs")

        for r in range(2, ws.max_row + 1):
            rama = _norm(ws.cell(r, col_rama).value)
            if rama.lower() not in ["funebres", "fúnebres"]:
                continue

            concepto_raw = _norm(ws.cell(r, col_concepto).value)
            mes_k = _mes_to_key(ws.cell(r, col_mes).value)
            if not mes_k or not concepto_raw:
                continue

            tipo_raw = _norm(ws.cell(r, col_tipo).value).lower() if col_tipo else ""
            monto_val = _to_float(ws.cell(r, col_monto).value) if col_monto else 0.0
            pct_val = _to_float(ws.cell(r, col_pct).value) if col_pct else 0.0
            obs_raw = _norm(ws.cell(r, col_obs).value) if col_obs else ""

            # Determinar tipo
            tipo = "pct" if ("por" in tipo_raw or "%" in tipo_raw) else "monto"

            # Etiquetas amigables (como en el HTML offline)
            cl = concepto_raw.lower()
            label = concepto_raw
            if "indument" in cl:
                label = "Indumentaria"
            elif "general" in cl:
                # Ojo: este concepto suele venir como "... incluidos choferes", por eso
                # se evalúa ANTES que el de chofer/furgonero.
                label = "Resto del personal"
            elif "cadaver" in cl or "cadáver" in cl or "no incluido" in cl or "inciso" in cl:
                label = "Manipulación de cadáveres"
            elif "furgon" in cl or "chofer/furgon" in cl:
                label = "Chofer/Furgonero"

            funebres_adic.setdefault(mes_k, []).append({
                "id": concepto_raw,   # id estable (se usa en fun_adic[] del /calcular)
                "label": label,
                "tipo": tipo,
                "monto": monto_val if tipo == "monto" else 0.0,
                "pct": pct_val if tipo == "pct" else 0.0,
                "obs": obs_raw,
            })


    # ---------------------------
    # Build meta
    # ---------------------------
    ramas = sorted(ramas_set)
    meses = sorted(meses_set)

    agrupamientos: Dict[str, List[str]] = {}
    categorias: Dict[str, Dict[str, List[str]]] = {}

    for rama in ramas:
        agrupamientos[rama] = sorted(list(agrup_by_rama.get(rama, set())))
        categorias[rama] = {}
        for agr in agrupamientos[rama]:
            categorias[rama][agr] = sorted(list(cat_by_rama_agrup.get((rama, agr), set())))

    return {
        "payload": payload,
        "meta": {
            "ramas": ramas,
            "meses": meses,
            "agrupamientos": agrupamientos,
            "categorias": categorias,
        },
        "meses_by_rama": {k: sorted(list(v)) for k, v in meses_by_rama.items()},
        "funebres_adic": funebres_adic,
    }


# ---------------------------
# Public API (used by main.py)
# ---------------------------

def get_meta() -> Dict[str, Any]:
    return _build_index()["meta"]

def get_payload(
    rama: str,
    mes: str,
    agrup: str = "—",
    categoria: str = "—",
    conex_cat: str = "",
    conexiones: int = 0,
    fun_adic: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Devuelve los valores base del maestro para la combinación dada.

    Se usa en:
      - /payload (solo rama + mes)
      - /calcular (rama + mes + agrup + categoria) como base.
    """
    idx = _build_index()
    key = (_norm(rama).upper(), _norm(agrup) or "—", _norm(categoria) or "—", _mes_to_key(mes))
    rec = idx["payload"].get(key)

    if not rec:
        # fallback: algunos front mandan "—" en agrup/cat o vienen vacíos
        key2 = (_norm(rama).upper(), "—", "—", _mes_to_key(mes))
        rec = idx["payload"].get(key2)

    if not rec:
        return {
            "ok": False,
            "error": "No se encontró esa combinación en el maestro",
            "rama": _norm(rama).upper(),
            "agrup": _norm(agrup) or "—",
            "categoria": _norm(categoria) or "—",
            "mes": _mes_to_key(mes),
        }

    labels = _nr_labels(key[0])

    out = {"ok": True, "rama": key[0], "agrup": key[1], "categoria": key[2], "mes": key[3], **rec, "labels": labels}

    # Agua Potable: ajustar valores base según selector de conexiones (A/B/C/D)
    if norm_rama(key[0]) in ("AGUA POTABLE", "AGUA", "AGUAPOTABLE") and (conex_cat or conexiones):
        regla = match_regla_conexiones(conex_cat or conexiones)
        try:
            f = float(regla.get("factor", 1.0))
        except Exception:
            f = 1.0
        if f and f != 1.0:
            out["basico"] = _round2(out.get("basico", 0.0) * f)
            out["no_rem"] = _round2(out.get("no_rem", 0.0) * f)
            out["suma_fija"] = _round2(out.get("suma_fija", 0.0) * f)
            out["conex_cat"] = _norm(conex_cat).upper() if conex_cat else ""
            out["conexiones"] = int(conexiones or 0)

    return out


# Compat: algunos módulos históricos importan find_row().
# Devuelve el registro del maestro para (rama, agrup, categoria, mes) o None si no existe.
def find_row(rama: str, agrup: str, categoria: str, mes: str) -> Optional[Dict[str, Any]]:
    res = get_payload(rama=rama, mes=mes, agrup=agrup, categoria=categoria)
    if not isinstance(res, dict) or not res.get("ok"):
        return None
    return res

def calcular_payload(
    rama: str,
    agrup: str,
    categoria: str,
    mes: str,
    jornada: float = 48,
    anios_antig: float = 0,
    osecac: bool = True,
    afiliado: bool = False,
    sind_pct: float = 0,
    sind_fijo: float = 0,
    titulo_pct: float = 0,
    zona_pct: float = 0,
    fer_no_trab: int = 0,
    fer_trab: int = 0,
    vac_goz: int = 0,
    aus_inj: int = 0,
    # Etapa 8: Jubilado / Suspensión-Licencia sin goce / Embargo
    jubilado: bool = False,
    susp_dias: int = 0,
    embargo: float = 0,
    # Horas (cantidades)
    hex50: float = 0,
    hex100: float = 0,
    hs_noct: float = 0,

    # Adicional por KM (Art. 36 Chofer/Ayudante)
    km_tipo: str = "",
    km_menos100: float = 0,
    km_mas100: float = 0,
    # Etapa 5/6: A cuenta (REM) / Viáticos (NR sin aportes)
    a_cuenta_rem: float = 0,
    viaticos_nr: float = 0,

    # Etapa 7: Manejo de Caja / Vidriera / Adelanto
    manejo_caja: bool = False,
    cajero_tipo: str = "",
    faltante_caja: float = 0,
    armado_vidriera: bool = False,
    adelanto_sueldo: float = 0,
    # SAC (estimación proporcional en mensual)
    sac_prop_mes: bool = False,
    # Agua potable: selector A/B/C/D (impacta en básicos y NR). Se mantiene conexiones por compatibilidad.
    conex_cat: str = "",
    conexiones: int = 0,
    # Fúnebres: ids de adicionales seleccionados (coma-separados)
    fun_adic: str = "",
) -> Dict[str, Any]:
    """Cálculo del endpoint /calcular (servidor).

    El front NO calcula: solo renderiza.
    Devuelve items + totales numéricos para que el HTML muestre cada fila.

    Versión núcleo (GENERAL): Básico, Antigüedad, Presentismo, NR base y descuentos principales.
    """
    base = get_payload(rama=rama, mes=mes, agrup=agrup, categoria=categoria)
    if not base.get("ok"):
        return base

    # -------- Bases prorrateadas (48hs) --------
    # CALL CENTER: la categoría ya trae su jornada (20/21/24/30/34/35/36/48hs).
    # No se prorratea por selector (evita que el básico se achique al poner 20hs).
    is_call = norm_rama(rama) in ("CALL CENTER", "CALLCENTER", "CALL", "CENTRO DE LLAMADAS", "CENTRO DE LLAMADA")
    hs_cat = _extract_hs_from_categoria(categoria) if is_call else None

    if is_call and hs_cat:
        j = float(hs_cat)
        factor = 1.0
        call_to_48 = (48.0 / j) if j else 1.0
    else:
        j = float(jornada or 48)
        factor = (j / 48.0) if 48.0 else 1.0
        call_to_48 = 1.0


    bas_base = float(base.get("basico", 0.0) or 0.0)
    nr_base = float(base.get("no_rem", 0.0) or 0.0)
    sf_base = float(base.get("suma_fija", 0.0) or 0.0)

    # Agua Potable: Conexiones (A/B/C/D) NO se muestra como adicional;
    # modifica directamente el valor del Básico y de los No Rem.
    is_agua = norm_rama(rama) in ("AGUA POTABLE", "AGUA", "AGUAPOTABLE")
    if is_agua:
        nivel = _norm(conex_cat).upper() if conex_cat else ""
        info = match_regla_conexiones(nivel if nivel else conexiones)
        fac = float(info.get("factor", 1.0) or 1.0)
        if fac and fac != 1.0:
            bas_base *= fac
            nr_base *= fac
            sf_base *= fac

    bas = bas_base * factor
    nr = nr_base * factor
    sf = sf_base * factor

    # NR base total (sin derivados). Se usa también para valor-hora NR.
    nr_base_total = round2(nr + sf)

    # -------- Horas extra / nocturnas --------
    # Reglas: divisor fijo 200. Nocturnas = recargo 13,33% (1h = 1h08m).
    def _fh(x: float) -> float:
        try:
            v = float(x or 0.0)
        except Exception:
            v = 0.0
        return max(0.0, v)

    hex50_h = _fh(hex50)
    hex100_h = _fh(hex100)
    hs_noct_h = _fh(hs_noct)

    # -------- Bases NR --------
    nr_base_total = round2(nr + sf)

    # -------- Horas (extras + nocturnas) --------
    # Divisor fijo: 200 hs.
    def _h(x) -> float:
        try:
            return max(0.0, float(x or 0.0))
        except Exception:
            return 0.0

    hex50_h = _h(hex50)
    hex100_h = _h(hex100)
    hs_noct_h = _h(hs_noct)

    # -------- Adicional por KM (Art. 36) --------
    # Regla histórica (Acuerdo 26/09/1983):
    #  - Ayudante: 0,0082% (primeros 100km) sobre básico inicial Auxiliar A
    #             0,01%   (>100km)         sobre básico inicial Auxiliar Especializado A
    #  - Chofer:   0,01%   (primeros 100km) sobre básico inicial Auxiliar B
    #             0,0115% (>100km)         sobre básico inicial Auxiliar Especializado B
    #
    # El front manda:
    #   km_tipo: "AY" (Ayudante) o "CH" (Chofer)
    #   km_menos100: km dentro de los primeros 100
    #   km_mas100: km por encima de 100
    #
    # Se prorratea por jornada (factor) igual que el básico (salvo Call Center, donde factor=1).
    def _canon(s: str) -> str:
        s = _norm(s).upper()
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _basico_ref(_rama: str, _mes: str, candidates: List[str]) -> float:
        idx = _build_index()
        mes_k = _mes_to_key(_mes)
        cand_can = [_canon(c) for c in candidates]

        def _search(rama_k: str) -> float:
            # 1) match exacto
            for (r, _agr, cat, m), rec in idx.get("payload", {}).items():
                if r != rama_k or m != mes_k:
                    continue
                cat_c = _canon(cat)
                if "MENORES" in cat_c:
                    continue
                if cat_c in cand_can:
                    try:
                        return float(rec.get("basico") or 0.0)
                    except Exception:
                        return 0.0
            # 2) contiene
            for (r, _agr, cat, m), rec in idx.get("payload", {}).items():
                if r != rama_k or m != mes_k:
                    continue
                cat_c = _canon(cat)
                if "MENORES" in cat_c:
                    continue
                if any(cc in cat_c for cc in cand_can):
                    try:
                        return float(rec.get("basico") or 0.0)
                    except Exception:
                        return 0.0
            return 0.0

        r0 = _canon(_rama)
        v = _search(r0)
        if (not v) and r0 != "GENERAL":
            v = _search("GENERAL")
        return float(v or 0.0)

    km_tipo_n = _norm(km_tipo).upper()
    km_le100 = max(0.0, float(km_menos100 or 0.0))
    km_gt100 = max(0.0, float(km_mas100 or 0.0))

    # 2 renglones (<=100 y >100) para que el "Base" muestre el básico referencia correcto.
    km_rem_le = 0.0
    km_rem_gt = 0.0
    km_base_le = 0.0
    km_base_gt = 0.0

    # Turismo (CCT 547/08): adicionales por KM con valores fijos por categoría operativa (C4/C5)
    is_turismo = norm_rama(rama) == "TURISMO"
    tur_cat = None
    if is_turismo:
        if "C4" in km_tipo_n:
            tur_cat = "C4"
        elif "C5" in km_tipo_n:
            tur_cat = "C5"

    if is_turismo and tur_cat and (km_le100 or km_gt100):
        TUR_KM_RATES = {
            "C4": {
                "2026-01": {"le": 112.31, "gt": 129.16},
                "2026-02": {"le": 112.31, "gt": 129.16},
                "2026-03": {"le": 112.31, "gt": 129.16},
                "2026-04": {"le": 112.31, "gt": 129.16},
                "2026-05": {"le": 122.31, "gt": 140.66},
            },
            "C5": {
                "2026-01": {"le": 110.62, "gt": 127.21},
                "2026-02": {"le": 110.62, "gt": 127.21},
                "2026-03": {"le": 110.62, "gt": 127.21},
                "2026-04": {"le": 110.62, "gt": 127.21},
                "2026-05": {"le": 120.62, "gt": 138.71},
            },
        }

        def _pick_rate(rmap: Dict[str, Dict[str, float]], mes_k: str) -> Dict[str, float]:
            if not rmap:
                return {"le": 0.0, "gt": 0.0}
            keys = sorted(rmap.keys())
            chosen = keys[0]
            for k in keys:
                if k <= mes_k:
                    chosen = k
            return rmap.get(chosen) or {"le": 0.0, "gt": 0.0}

        mes_k = _mes_to_key(mes)
        rates = _pick_rate(TUR_KM_RATES.get(tur_cat, {}), mes_k)
        rate_le = float(rates.get("le") or 0.0)
        rate_gt = float(rates.get("gt") or 0.0)

        # En Turismo el "Base" lo mostramos como $/km (igual que en la escala).
        km_base_le = rate_le
        km_base_gt = rate_gt
        km_rem_le = round2(rate_le * km_le100) if (rate_le and km_le100) else 0.0
        km_rem_gt = round2(rate_gt * km_gt100) if (rate_gt and km_gt100) else 0.0

    elif km_tipo_n in ("AY", "AYUDANTE", "CH", "CHOFER") and (km_le100 or km_gt100):
        if km_tipo_n in ("AY", "AYUDANTE"):
            km_base_le = _basico_ref(rama, mes, ["AUXILIAR A", "AUXILIAR  A", "PERSONAL AUXILIAR A", "AUXILIAR LETRA A"])
            km_base_gt = _basico_ref(rama, mes, ["AUXILIAR ESPECIALIZADO A", "AUXILIAR  ESPECIALIZADO A"])
            # Art. 36: adicional por km recorrido (no se prorratea por jornada).
            km_rem_le = round2(km_base_le * 0.000082 * km_le100) if (km_base_le and km_le100) else 0.0
            km_rem_gt = round2(km_base_gt * 0.0001 * km_gt100) if (km_base_gt and km_gt100) else 0.0
        else:
            km_base_le = _basico_ref(rama, mes, ["AUXILIAR B", "AUXILIAR  B", "PERSONAL AUXILIAR B", "AUXILIAR LETRA B"])
            km_base_gt = _basico_ref(rama, mes, ["AUXILIAR ESPECIALIZADO B", "AUXILIAR  ESPECIALIZADO B"])
            # Art. 36: adicional por km recorrido (no se prorratea por jornada).
            km_rem_le = round2(km_base_le * 0.0001 * km_le100) if (km_base_le and km_le100) else 0.0
            km_rem_gt = round2(km_base_gt * 0.000115 * km_gt100) if (km_base_gt and km_gt100) else 0.0

    km_rem_total = round2(km_rem_le + km_rem_gt)

    DIV_HORA = 200.0
    hora_rem = (float(bas) / DIV_HORA) if bas else 0.0
    hora_nr = (float(nr_base_total) / DIV_HORA) if nr_base_total else 0.0

    hex50_rem = round2(hora_rem * 1.5 * hex50_h) if (hora_rem and hex50_h) else 0.0
    hex50_nr = round2(hora_nr * 1.5 * hex50_h) if (hora_nr and hex50_h) else 0.0
    hex100_rem = round2(hora_rem * 2.0 * hex100_h) if (hora_rem and hex100_h) else 0.0
    hex100_nr = round2(hora_nr * 2.0 * hex100_h) if (hora_nr and hex100_h) else 0.0

    # Hora nocturna: recargo 13,33% (1h nocturna = 1h 8m). Se liquida como adicional.
    NOCT_ADIC_PCT = 0.13333333333333333
    noct_rem = round2(hora_rem * NOCT_ADIC_PCT * hs_noct_h) if (hora_rem and hs_noct_h) else 0.0
    noct_nr = round2(hora_nr * NOCT_ADIC_PCT * hs_noct_h) if (hora_nr and hs_noct_h) else 0.0

    # -------- Cálculos núcleo --------
    # Remunerativos
    pct_ant = float(anios_antig or 0.0) * 0.01

    # Etapa 5/6: A cuenta (REM) / Viáticos (NR sin aportes)
    def _fpos(x) -> float:
        try:
            return max(0.0, float(x or 0.0))
        except Exception:
            return 0.0

    a_cuenta = _fpos(a_cuenta_rem)
    viaticos = _fpos(viaticos_nr)

    # Etapa 7: Manejo de Caja / Vidriera / Adelanto / Faltante
    # - Manejo de Caja (REM): A/C 12,25% sobre básico inicial Cajero A; B 48% sobre básico inicial Cajero B.
    #   Regla del sistema: el adicional se considera ANUAL, por lo que se liquida mensualmente como (importe anual / 12).
    # - Armado de vidriera (REM): 3,83% sobre básico inicial Vendedor B.
    # - Adelanto de sueldo y Faltante de caja: descuentos (no afectan bases de aportes).

    caj_tipo = str(cajero_tipo or "").strip().upper()
    manejo_caja_ok = bool(manejo_caja) and caj_tipo in ("A", "B", "C")

    caja_base = 0.0
    caja_pct = 0.0
    if manejo_caja_ok:
        if caj_tipo in ("A", "C"):
            caja_base = _basico_ref(rama, mes, ["CAJERO A", "CAJEROS A", "CAJERO  A", "CAJERO A "])
            caja_pct = 0.1225
        elif caj_tipo == "B":
            caja_base = _basico_ref(rama, mes, ["CAJERO B", "CAJEROS B", "CAJERO  B", "CAJERO B "])
            caja_pct = 0.48

    # ANUAL -> mensual (/12). Se prorratea por jornada usando factor (j/48).
    caja_rem = round2((caja_base * caja_pct * factor) / 12.0) if (caja_base and caja_pct) else 0.0
    caja_rem_os = round2((caja_base * caja_pct) / 12.0) if (caja_base and caja_pct) else 0.0

    vid_base = _basico_ref(rama, mes, ["VENDEDOR B", "Vendedor B", "VENDEDOR  B"]) if bool(armado_vidriera) else 0.0
    vid_pct = 0.0383
    vid_rem = round2(vid_base * vid_pct * factor) if (vid_base and bool(armado_vidriera)) else 0.0
    vid_rem_os = round2(vid_base * vid_pct) if (vid_base and bool(armado_vidriera)) else 0.0

    faltante = _fpos(faltante_caja)
    adelanto = _fpos(adelanto_sueldo)
    # El faltante se descuenta ÚNICAMENTE hasta el monto del adicional de Manejo de Caja.
    faltante_desc = round2(min(faltante, caja_rem)) if (faltante and caja_rem) else 0.0

    # Zona desfavorable (porcentaje sobre Básico prorrateado)
    try:
        zona_pct_f = float(zona_pct or 0.0)
    except Exception:
        zona_pct_f = 0.0
    zona_pct_f = max(0.0, zona_pct_f)
    zona = round2(bas * (zona_pct_f / 100.0)) if zona_pct_f else 0.0

    # Antigüedad: base incluye Zona (criterio del sistema para cálculos generales)
    base_ant = round2(bas + zona)
    antig = round2(base_ant * pct_ant)

    # Regla Presentismo: se pierde con 2 (dos) o más ausencias injustificadas.
    aus_dias = max(0, int(aus_inj or 0))
    presentismo_habil = (aus_dias < 2)
    # Presentismo: doceava parte de (Básico + Zona + Antigüedad + Horas + Adicionales)
    # Incluye: horas extra/nocturnas, adicional por KM y A cuenta (REM).
    base_pres = round2(bas + zona + antig + hex50_rem + hex100_rem + noct_rem + km_rem_total + caja_rem + vid_rem + a_cuenta)
    presentismo = round2(base_pres / 12.0) if presentismo_habil else 0.0

    rem_total = round2(bas + zona + presentismo + antig + hex50_rem + hex100_rem + noct_rem + km_rem_total + caja_rem + vid_rem + a_cuenta)

    # No remunerativos (NR) + derivados (Antigüedad NR / Presentismo NR)
    antig_nr = round2(nr_base_total * pct_ant) if nr_base_total else 0.0
    # Presentismo sobre NR: misma lógica que REM (12ava parte), incluyendo Antigüedad NR
    # y también horas extra/nocturnas NR.
    # Se pierde si hay 2+ ausencias injustificadas.
    base_pres_nr = round2(nr_base_total + antig_nr + hex50_nr + hex100_nr + noct_nr)
    presentismo_nr = round2(base_pres_nr / 12.0) if (base_pres_nr and presentismo_habil) else 0.0

    nr_total = round2(nr_base_total + antig_nr + presentismo_nr + hex50_nr + hex100_nr + noct_nr)

    # Viáticos (NR sin aportes): se suman al NR a pagar, pero NO integran bases de aportes
    # ni el Presentismo sobre NR.
    if viaticos:
        nr_total = round2(nr_total + viaticos)

    # -------- FUNEBRES: Adicionales (según maestro) --------
    fun_rows: List[Dict[str, Any]] = []
    if norm_rama(base["rama"]) in ("FUNEBRES", "FÚNEBRES"):
        sel_raw = (fun_adic or "").strip()
        if sel_raw:
            # IMPORTANTE: NO cortar por coma, porque algunos IDs contienen comas
            # (p.ej. "incluidos choferes"). Usamos solo ";" como separador.
            sel_ids = [s.strip() for s in sel_raw.split(";") if s.strip()]
            if sel_ids:
                defs = get_adicionales_funebres(mes)
                by_id = {str(d.get("id")): d for d in defs}
                for sid in sel_ids:
                    d = by_id.get(str(sid))
                    if not d:
                        continue
                    label = str(d.get("label") or sid)
                    tipo = str(d.get("tipo") or "").strip().lower()
                    monto = float(d.get("monto") or 0.0)
                    pct = float(d.get("pct") or 0.0)

                    val = 0.0
                    base_num = 0.0
                    if tipo in ("monto", "importe", "fijo") and monto:
                        # prorrateo por jornada
                        val = round2(monto * factor)
                    elif pct:
                        base_num = float(bas)
                        val = round2(bas * (pct / 100.0))
                    elif monto:
                        val = round2(monto * factor)

                    if val:
                        fun_rows.append({"label": label, "val": float(val), "base": float(base_num)})
                        rem_total = round2(rem_total + val)

    # -------- TURISMO: Adicional por Título --------
    # Se aplica sobre el básico (REM) y sobre el NR de $40.000 (en nuestro maestro = suma_fija).
    try:
        titulo_pct_f = float(titulo_pct or 0.0)
    except Exception:
        titulo_pct_f = 0.0
    titulo_pct_f = max(0.0, titulo_pct_f)

    titulo_rem = 0.0
    titulo_nr = 0.0
    if base["rama"] == "TURISMO" and titulo_pct_f > 0:
        titulo_rem = round2(bas * (titulo_pct_f / 100.0)) if bas else 0.0
        # Turismo: $40.000 NR corresponde a suma_fija (sf)
        titulo_nr = round2(sf * (titulo_pct_f / 100.0)) if sf else 0.0
        rem_total = round2(rem_total + titulo_rem)
        nr_total = round2(nr_total + titulo_nr)

    # -------- Feriados --------
    fer_no = max(0, int(fer_no_trab or 0))
    fer_si = max(0, int(fer_trab or 0))
    base_fer_rem = round2(bas + zona + antig)
    base_fer_nr = round2(nr_base_total + antig_nr)
    # Para mensualizados:
    # - Feriado NO trabajado: se suma la diferencia entre día feriado (1/25) y día normal incluido en el mensual (1/30).
    # - Feriado trabajado: se suma 1 día feriado (1/25).
    vdia25_rem = round2(base_fer_rem / 25.0) if base_fer_rem else 0.0
    vdia30_rem = round2(base_fer_rem / 30.0) if base_fer_rem else 0.0
    vdia25_nr = round2(base_fer_nr / 25.0) if base_fer_nr else 0.0
    vdia30_nr = round2(base_fer_nr / 30.0) if base_fer_nr else 0.0

    fer_no_rem = round2(fer_no * (vdia25_rem - vdia30_rem)) if fer_no else 0.0
    fer_si_rem = round2(fer_si * vdia25_rem) if fer_si else 0.0

    fer_no_nr = round2(fer_no * (vdia25_nr - vdia30_nr)) if fer_no else 0.0
    fer_si_nr = round2(fer_si * vdia25_nr) if fer_si else 0.0

    rem_total = round2(rem_total + fer_no_rem + fer_si_rem)
    nr_total = round2(nr_total + fer_no_nr + fer_si_nr)

    # -------- Vacaciones gozadas --------
    # Para mensualizados: plus por divisor 1/25 vs día normal (1/30).
    vac_goz_dias = max(0, int(vac_goz or 0))
    vac_goz_rem = 0.0
    vac_goz_nr = 0.0
    if vac_goz_dias:
        vac_goz_rem = round2(vac_goz_dias * (vdia25_rem - vdia30_rem))
        vac_goz_nr = round2(vac_goz_dias * (vdia25_nr - vdia30_nr))
        rem_total = round2(rem_total + vac_goz_rem)
        nr_total = round2(nr_total + vac_goz_nr)

    # -------- SAC (Jun/Dic) o SAC proporcional (mes) --------
    sac_concepto = ""
    sac_row_rem = 0.0
    sac_row_nr = 0.0
    sac_row_base = 0.0
    mes_num = 0
    try:
        mes_num = int(str(base.get("mes") or mes or "").split("-")[1])
    except Exception:
        mes_num = 0

    # Base SAC = Base mensual + Presentismo (REM y NR), sin extras/vacaciones.
    base_sac_rem = round2((bas + zona + antig) + (presentismo if presentismo_habil else 0.0))
    base_sac_nr = round2((nr_base_total + antig_nr) + (presentismo_nr if presentismo_habil else 0.0))
    sac_row_base = round2(base_sac_rem + base_sac_nr)

    if mes_num in (6, 12):
        sac_concepto = "SAC (Junio)" if mes_num == 6 else "SAC (Diciembre)"
        sac_row_rem = round2(base_sac_rem * 0.5)
        sac_row_nr = round2(base_sac_nr * 0.5)
        rem_total = round2(rem_total + sac_row_rem)
        nr_total = round2(nr_total + sac_row_nr)
    elif bool(sac_prop_mes) and (1 <= mes_num <= 12):
        # Estimación: Base del mes * (meses del semestre / 12)
        meses_sem = mes_num if mes_num <= 6 else (mes_num - 6)
        factor_sac = float(meses_sem) / 12.0
        sac_concepto = "SAC proporcional (mes)"
        sac_row_rem = round2(base_sac_rem * factor_sac)
        sac_row_nr = round2(base_sac_nr * factor_sac)
        rem_total = round2(rem_total + sac_row_rem)
        nr_total = round2(nr_total + sac_row_nr)

    # -------- Ausencias injustificadas (descuento) --------
    base_dia_aus = round2((bas + zona + antig) / 30.0) if (bas or zona or antig) else 0.0
    aus_rem = round2(aus_dias * base_dia_aus) if aus_dias else 0.0

    # -------- Suspensión / Licencia sin goce (descuento) --------
    susp_d = max(0, int(susp_dias or 0))
    base_dia_susp = base_dia_aus  # mismo criterio que ausencias (Básico+Zona+Antig) / 30
    susp_rem = round2(susp_d * base_dia_susp) if susp_d else 0.0

    # Base imponible para aportes: REM - ausencias - suspensión/LSG
    rem_aportes = max(0.0, round2(rem_total - aus_rem - susp_rem))

    # Descuentos
    # Regla especial (definida por el admin): cuando el trabajador es JUBILADO,
    # los descuentos de ley quedan:
    #   - Jubilación 11%
    #   - FAECYS 0,5%
    #   - Sindicato 2%
    #   - Afiliación 2% (solo si está marcado Afiliado)
    # y NO se descuenta PAMI/OS/OSECAC.
    jub = round2(rem_aportes * 0.11)
    pami = 0.0 if bool(jubilado) else round2(rem_aportes * 0.03)
    # Obra Social (OSECAC): BASE JORNADA COMPLETA (48hs), sin prorrateo por jornada.
    # Importante: no "desprorrateamos" totales, porque eso infla importes fijos (p.ej. a-cuenta).
    # Recalculamos una simulación a 48hs manteniendo el resto de parámetros (antig., zona, feriados, ausencias, etc.).
    bas_os = float(bas_base) * call_to_48  # 48hs (CALL: simula a 48). Agua: ya incluye conexiones en bas_base.
    nr_os = float(nr_base) * call_to_48
    sf_os = float(sf_base) * call_to_48

    zona_os = round2(bas_os * (zona_pct_f / 100.0)) if zona_pct_f else 0.0
    base_ant_os = round2(bas_os + zona_os)
    antig_os = round2(base_ant_os * pct_ant)
    # Horas (48hs) – mismo input de horas, con valor hora simulado a 48hs
    hora_rem_os = (float(bas_os) / DIV_HORA) if bas_os else 0.0
    # OJO: para NR, la base hora es (nr_os + sf_os)
    nr_base_total_os = round2(nr_os + sf_os)
    hora_nr_os = (float(nr_base_total_os) / DIV_HORA) if nr_base_total_os else 0.0
    hex50_rem_os = round2(hora_rem_os * 1.5 * hex50_h) if (hora_rem_os and hex50_h) else 0.0
    hex50_nr_os = round2(hora_nr_os * 1.5 * hex50_h) if (hora_nr_os and hex50_h) else 0.0
    hex100_rem_os = round2(hora_rem_os * 2.0 * hex100_h) if (hora_rem_os and hex100_h) else 0.0
    hex100_nr_os = round2(hora_nr_os * 2.0 * hex100_h) if (hora_nr_os and hex100_h) else 0.0
    noct_rem_os = round2(hora_rem_os * NOCT_ADIC_PCT * hs_noct_h) if (hora_rem_os and hs_noct_h) else 0.0
    noct_nr_os = round2(hora_nr_os * NOCT_ADIC_PCT * hs_noct_h) if (hora_nr_os and hs_noct_h) else 0.0

    # Incluye A cuenta (REM) como monto fijo (no se prorratea por la simulación a 48hs).
    base_pres_os = round2(bas_os + zona_os + antig_os + hex50_rem_os + hex100_rem_os + noct_rem_os + km_rem_total + caja_rem_os + vid_rem_os + a_cuenta)
    presentismo_os = round2(base_pres_os / 12.0) if presentismo_habil else 0.0
    rem_total_os = round2(bas_os + zona_os + antig_os + presentismo_os + hex50_rem_os + hex100_rem_os + noct_rem_os + km_rem_total + caja_rem_os + vid_rem_os + a_cuenta)

    antig_nr_os = round2(nr_base_total_os * pct_ant) if nr_base_total_os else 0.0
    presentismo_nr_os = (
        round2((nr_base_total_os + antig_nr_os + hex50_nr_os + hex100_nr_os + noct_nr_os) / 12.0)
        if (nr_base_total_os and presentismo_habil)
        else 0.0
    )
    nr_total_os = round2(nr_base_total_os + antig_nr_os + presentismo_nr_os + hex50_nr_os + hex100_nr_os + noct_nr_os)

    # FUNEBRES: adicionales (48hs)
    if norm_rama(base["rama"]) in ("FUNEBRES", "FÚNEBRES"):
        sel_raw = (fun_adic or "").strip()
        if sel_raw:
            sel_ids = [s.strip() for s in sel_raw.split(";") if s.strip()]
            if sel_ids:
                defs = get_adicionales_funebres(mes)
                by_id = {str(d.get("id")): d for d in defs}
                for sid in sel_ids:
                    d = by_id.get(str(sid))
                    if not d:
                        continue
                    tipo = str(d.get("tipo") or "").strip().lower()
                    monto = float(d.get("monto") or 0.0)
                    pct = float(d.get("pct") or 0.0)
                    val = 0.0
                    if tipo in ("monto", "importe", "fijo") and monto:
                        val = round2(monto)  # 48hs
                    elif pct:
                        val = round2(bas_os * (pct / 100.0))
                    elif monto:
                        val = round2(monto)
                    if val:
                        rem_total_os = round2(rem_total_os + val)

    # TURISMO: adicional por título (48hs)
    if base["rama"] == "TURISMO" and titulo_pct_f > 0:
        titulo_rem_os = round2(bas_os * (titulo_pct_f / 100.0)) if bas_os else 0.0
        titulo_nr_os = round2(sf_os * (titulo_pct_f / 100.0)) if sf_os else 0.0
        rem_total_os = round2(rem_total_os + titulo_rem_os)
        nr_total_os = round2(nr_total_os + titulo_nr_os)

    # Feriados (48hs)
    base_fer_rem_os = round2(bas_os + zona_os + antig_os)
    base_fer_nr_os = round2(nr_base_total_os + antig_nr_os)
    vdia25_rem_os = round2(base_fer_rem_os / 25.0) if base_fer_rem_os else 0.0
    vdia30_rem_os = round2(base_fer_rem_os / 30.0) if base_fer_rem_os else 0.0
    vdia25_nr_os = round2(base_fer_nr_os / 25.0) if base_fer_nr_os else 0.0
    vdia30_nr_os = round2(base_fer_nr_os / 30.0) if base_fer_nr_os else 0.0

    fer_no_rem_os = round2(fer_no * (vdia25_rem_os - vdia30_rem_os)) if fer_no else 0.0
    fer_si_rem_os = round2(fer_si * vdia25_rem_os) if fer_si else 0.0
    fer_no_nr_os = round2(fer_no * (vdia25_nr_os - vdia30_nr_os)) if fer_no else 0.0
    fer_si_nr_os = round2(fer_si * vdia25_nr_os) if fer_si else 0.0

    rem_total_os = round2(rem_total_os + fer_no_rem_os + fer_si_rem_os)
    nr_total_os = round2(nr_total_os + fer_no_nr_os + fer_si_nr_os)

    # Vacaciones gozadas: plus divisor... (mismo criterio, pero sobre base OS)
    if vac_goz_dias:
        vac_goz_rem_os = round2(vac_goz_dias * (vdia25_rem_os - vdia30_rem_os))
        vac_goz_nr_os = round2(vac_goz_dias * (vdia25_nr_os - vdia30_nr_os))
        rem_total_os = round2(rem_total_os + vac_goz_rem_os)
        nr_total_os = round2(nr_total_os + vac_goz_nr_os)

    # SAC (48hs para base de Obra Social)
    if mes_num in (6, 12):
        base_sac_rem_os = round2((bas_os + zona_os + antig_os) + (presentismo_os if presentismo_habil else 0.0))
        base_sac_nr_os = round2((nr_base_total_os + antig_nr_os) + (presentismo_nr_os if presentismo_habil else 0.0))
        rem_total_os = round2(rem_total_os + round2(base_sac_rem_os * 0.5))
        nr_total_os = round2(nr_total_os + round2(base_sac_nr_os * 0.5))
    elif bool(sac_prop_mes) and (1 <= mes_num <= 12):
        meses_sem = mes_num if mes_num <= 6 else (mes_num - 6)
        factor_sac = float(meses_sem) / 12.0
        base_sac_rem_os = round2((bas_os + zona_os + antig_os) + (presentismo_os if presentismo_habil else 0.0))
        base_sac_nr_os = round2((nr_base_total_os + antig_nr_os) + (presentismo_nr_os if presentismo_habil else 0.0))
        rem_total_os = round2(rem_total_os + round2(base_sac_rem_os * factor_sac))
        nr_total_os = round2(nr_total_os + round2(base_sac_nr_os * factor_sac))

    # Ausencias (48hs)
    base_dia_aus_os = round2((bas_os + zona_os + antig_os) / 30.0) if (bas_os or zona_os or antig_os) else 0.0
    aus_rem_os = round2(aus_dias * base_dia_aus_os) if aus_dias else 0.0
    susp_rem_os = round2(susp_d * base_dia_aus_os) if susp_d else 0.0
    rem_aportes_os = max(0.0, round2(rem_total_os - aus_rem_os - susp_rem_os))

    # Obra social y aporte fijo: para JUBILADO se anulan, aun si está tildado OSECAC.
    if bool(jubilado):
        os_base = round2(rem_aportes_os + nr_total_os)
        os_aporte = 0.0
        osecac_100 = 0.0
    else:
        os_base = round2((rem_aportes_os + nr_total_os) if bool(osecac) else rem_aportes_os)
        os_aporte = round2(os_base * 0.03) if bool(osecac) else 0.0
        osecac_100 = 100.0 if bool(osecac) else 0.0

    # Base para aportes porcentuales (Sindicato/FAECYS, etc.): excluye viáticos NR sin aportes.
    nr_aportable_real = max(0.0, round2(nr_total - (viaticos or 0.0)))

    # Aportes sindicales/FAECYS: base = REM aportable + NR aportable (excluye viáticos).
    # Para JUBILADO, el set de descuentos se reduce a: Jub 11% + FAECYS 0,5% + Sindicato 2% (+ Afiliación 2% si aplica).
    base_fs = 0.0
    faecys = 0.0
    sind_solid = 0.0
    sind_af = 0.0

    sind = 0.0
    sind_fijo_monto = 0.0

    if bool(jubilado):
        base_fs = round2(rem_aportes + nr_aportable_real)
        faecys = round2(base_fs * 0.005)
        sind_solid = round2(base_fs * 0.02)
        # En JUBILADO, la afiliación respeta el selector (% 1–4) y/o monto fijo.
        sind_af = 0.0

        if bool(afiliado):
            try:
                sp = float(sind_pct or 0.0)
            except Exception:
                sp = 0.0
            if sp > 0:
                sind = round2(base_fs * (sp / 100.0))

            try:
                sind_fijo_monto = round2(max(0.0, float(sind_fijo or 0.0)))
            except Exception:
                sind_fijo_monto = 0.0
    else:
        if bool(afiliado):
            try:
                sp = float(sind_pct or 0.0)
            except Exception:
                sp = 0.0
            if sp > 0:
                sind = round2((rem_aportes + nr_aportable_real) * (sp / 100.0))

            # Monto fijo de sindicato (se aplica SOLO si está afiliado).
            try:
                sind_fijo_monto = round2(max(0.0, float(sind_fijo or 0.0)))
            except Exception:
                sind_fijo_monto = 0.0

    ded_pre = round2(
        jub
        + pami
        + os_aporte
        + osecac_100
        + faecys
        + sind_solid
        + sind_af
        + sind
        + sind_fijo_monto
        + aus_rem
        + susp_rem
        + faltante_desc
        + adelanto
    )
    neto_pre = round2((rem_total + nr_total) - ded_pre)
    emb_in = 0.0
    try:
        emb_in = float(embargo or 0.0)
    except Exception:
        emb_in = 0.0
    emb_in = max(0.0, emb_in)
    embargo_monto = round2(min(emb_in, max(0.0, neto_pre))) if emb_in else 0.0
    ded_total = round2(ded_pre + embargo_monto)
    neto = round2(neto_pre - embargo_monto)

    def item(concepto: str, r: float = 0.0, n: float = 0.0, d: float = 0.0, base_num: float = 0.0) -> Dict[str, Any]:
        out = {"concepto": concepto, "r": float(r), "n": float(n), "d": float(d)}
        if base_num:
            out["base"] = float(base_num)
        return out

    items: List[Dict[str, Any]] = [item("Básico", r=bas, base_num=bas)]

    if zona:
        items.append(item("Zona desfavorable", r=zona, base_num=bas))

    if antig:
        items.append(item("Antigüedad", r=antig, base_num=base_ant))

    # Horas extra / nocturnas (2 filas o 4 si hay NR)
    if hex50_rem:
        items.append(item("Horas extra 50% (Rem)", r=hex50_rem, base_num=hora_rem))
    if hex50_nr:
        items.append(item("Horas extra 50% (NR)", n=hex50_nr, base_num=hora_nr))
    if hex100_rem:
        items.append(item("Horas extra 100% (Rem)", r=hex100_rem, base_num=hora_rem))
    if hex100_nr:
        items.append(item("Horas extra 100% (NR)", n=hex100_nr, base_num=hora_nr))
    if noct_rem:
        items.append(item("Horas nocturnas (Rem)", r=noct_rem, base_num=hora_rem))
    if noct_nr:
        items.append(item("Horas nocturnas (NR)", n=noct_nr, base_num=hora_nr))

    # Adicional por KM — 2 filas (<=100 / >100)
    if km_rem_le:
        if is_turismo and tur_cat:
            items.append(item(f'Adicional por KM (Turismo - Operativo {tur_cat}) ≤100 km ({km_le100:g} km)', r=km_rem_le, base_num=km_base_le))
        else:
            tipo_txt = 'Ayudante' if km_tipo_n in ('AY','AYUDANTE') else 'Chofer'
            items.append(item(f'Adicional por KM (Art. 36 - {tipo_txt}) ≤100 km ({km_le100:g} km)', r=km_rem_le, base_num=km_base_le))
    if km_rem_gt:
        if is_turismo and tur_cat:
            items.append(item(f'Adicional por KM (Turismo - Operativo {tur_cat}) >100 km ({km_gt100:g} km)', r=km_rem_gt, base_num=km_base_gt))
        else:
            tipo_txt = 'Ayudante' if km_tipo_n in ('AY','AYUDANTE') else 'Chofer'
            items.append(item(f'Adicional por KM (Art. 36 - {tipo_txt}) >100 km ({km_gt100:g} km)', r=km_rem_gt, base_num=km_base_gt))

    # Manejo de Caja (REM)
    if caja_rem:
        lbl_tipo = "Cajero B" if caj_tipo == "B" else "Cajero A/C"
        lbl_pct = "48%" if caj_tipo == "B" else "12,25%"
        items.append(item(
            f"Manejo de Caja ({lbl_tipo}) {lbl_pct}",
            r=caja_rem,
            base_num=caja_base,
        ))

    # Armado de vidriera (REM)
    if vid_rem:
        items.append(item(
            "Armado de vidriera (Art. 23) 3,83%",
            r=vid_rem,
            base_num=vid_base,
        ))

    # A cuenta futuros aumentos (REM)
    if a_cuenta:
        items.append(item(
            "A cuenta futuros aumentos (REM)",
            r=a_cuenta,
        ))

    # Presentismo: si se pierde por 2+ ausencias injustificadas, NO se muestra la fila (pedido César).
    if presentismo_habil and presentismo:
        items.append(item(
            "Presentismo",
            r=presentismo,
            base_num=base_pres,
        ))

    # SAC (Jun/Dic) o SAC proporcional (mes)
    if sac_concepto and (sac_row_rem or sac_row_nr):
        items.append(item(
            sac_concepto,
            r=sac_row_rem,
            n=sac_row_nr,
            base_num=sac_row_base,
        ))

    # Fúnebres: adicionales seleccionados (según maestro)
    if fun_rows:
        for fr in fun_rows:
            items.append(item(
                str(fr.get("label") or "Adicional"),
                r=float(fr.get("val") or 0.0),
                base_num=float(fr.get("base") or 0.0),
            ))



    # -------- Etapa 12: Mostrar adicionales (Turismo Título / Agua Conexiones) --------
    if titulo_rem:
        items.append(item(
            'Adicional por Título',
            r=titulo_rem,
            base_num=bas,
        ))
    if titulo_nr:
        items.append(item(
            'Adicional por Título (NR)',
            n=titulo_nr,
            base_num=sf,
        ))

    # Conexiones (Agua Potable): no se agrega fila, porque el selector modifica el básico y los NR.
    # Feriados (REM)
    if fer_no_rem:
        items.append(item(
            f"Feriado no trabajado ({fer_no} día{'s' if fer_no != 1 else ''})",
            r=fer_no_rem,
            base_num=base_fer_rem,
        ))
    if fer_si_rem:
        items.append(item(
            f"Feriado trabajado ({fer_si} día{'s' if fer_si != 1 else ''})",
            r=fer_si_rem,
            base_num=base_fer_rem,
        ))

    # Vacaciones gozadas: plus por divisor 1/25 vs 1/30
    if vac_goz_dias and vac_goz_rem:
        items.append(item(
            f"Vacaciones gozadas (plus 1/25) ({vac_goz_dias} día{'s' if vac_goz_dias != 1 else ''})",
            r=vac_goz_rem,
            base_num=base_fer_rem,
        ))

    labels = _nr_labels(base["rama"])

    if nr:
        items.append(item(labels.get("no_rem", "No Remunerativo"), n=nr))
    if sf:
        items.append(item(labels.get("suma_fija", "Suma Fija (NR)"), n=sf))

    # Viáticos (NR sin aportes)
    if viaticos:
        items.append(item(
            "Viáticos (NR sin aportes)",
            n=viaticos,
        ))

    # Derivados sobre NR (desglosado como filas NR)
    if antig_nr:
        items.append(item("Antigüedad (NR)", n=antig_nr, base_num=nr_base_total))
    # Presentismo sobre NR: si se pierde por 2+ ausencias injustificadas, NO se muestra la fila.
    if presentismo_habil and presentismo_nr:
        items.append(item(
            "Presentismo (NR)",
            n=presentismo_nr,
            base_num=base_pres_nr,
        ))

    # Feriados (NR)
    if fer_no_nr:
        items.append(item(
            f"Feriado no trabajado (NR) ({fer_no} día{'s' if fer_no != 1 else ''})",
            n=fer_no_nr,
            base_num=base_fer_nr,
        ))
    if fer_si_nr:
        items.append(item(
            f"Feriado trabajado (NR) ({fer_si} día{'s' if fer_si != 1 else ''})",
            n=fer_si_nr,
            base_num=base_fer_nr,
        ))

    if vac_goz_dias and vac_goz_nr:
        items.append(item(
            f"Vacaciones gozadas (NR) (plus 1/25) ({vac_goz_dias} día{'s' if vac_goz_dias != 1 else ''})",
            n=vac_goz_nr,
            base_num=base_fer_nr,
        ))

    # Ausencias injustificadas (descuento) – reduce bases de aportes vía rem_aportes
    if aus_rem:
        items.append(item(
            f"Ausencias injustificadas ({aus_dias} día{'s' if aus_dias != 1 else ''})",
            d=aus_rem,
            base_num=base_dia_aus,
        ))

    # Suspensión / Licencia sin goce (descuento)
    if susp_rem:
        items.append(item(
            f"Licencia sin goce / suspensión ({susp_d} día{'s' if susp_d != 1 else ''})",
            d=susp_rem,
            base_num=base_dia_susp,
        ))

    if faltante_desc:
        items.append(item(
            "Faltante de caja (desc.)",
            d=faltante_desc,
            base_num=caja_rem,
        ))

    if adelanto:
        items.append(item(
            "Adelanto de sueldo",
            d=adelanto,
        ))

    if embargo_monto:
        items.append(item(
            "Embargo (desc.)",
            d=embargo_monto,
            base_num=neto_pre,
        ))

    if bool(jubilado):
        items.append(item("Jubilación 11% (Jubilado)", d=jub, base_num=rem_aportes))
        items.append(item("FAECYS 0,5%", d=faecys, base_num=base_fs))
        items.append(item("Sindicato 2%", d=sind_solid, base_num=base_fs))
        if sind:
            items.append(item(f"Afiliación {float(sind_pct):g}%", d=sind, base_num=base_fs))
        if sind_fijo_monto:
            items.append(item("Afiliación (fijo)", d=sind_fijo_monto))
    else:
        items.append(item("Jubilación 11%", d=jub, base_num=rem_aportes))
        items.append(item("Ley 19.032 (PAMI) 3%", d=pami, base_num=rem_aportes))

        if bool(osecac):
            items.append(item("Obra Social 3%", d=os_aporte, base_num=os_base))
            items.append(item("OSECAC $100", d=osecac_100))
        else:
            items.append(item("Obra Social 3%", d=0.0, base_num=os_base))

        if sind:
            items.append(item(f"Sindicato {float(sind_pct):g}%", d=sind, base_num=(rem_aportes + nr_aportable_real)))
        if sind_fijo_monto:
            items.append(item("Sindicato (fijo)", d=sind_fijo_monto))

    return {
        "ok": True,
        "rama": base["rama"],
        "agrup": base["agrup"],
        "categoria": base["categoria"],
        "mes": base["mes"],
        "jornada": j,
        "anios_antig": float(anios_antig or 0),
        "osecac": bool(osecac),
        "afiliado": bool(afiliado),
        "sind_pct": float(sind_pct or 0),
        "sind_fijo": float(sind_fijo or 0),
        "titulo_pct": float(titulo_pct or 0),
        "zona_pct": float(zona_pct_f),

        "labels": labels,

        "basico_base": float(bas_base),
        "no_rem_base": float(nr_base),
        "suma_fija_base": float(sf_base),

        "basico": float(bas),
        "zona": float(zona),
        "no_rem": float(nr),
        "suma_fija": float(sf),

        "items": items,
        "totales": {
            "rem": float(rem_total),
            "nr": float(nr_total),
            "ded": float(ded_total),
            "neto": float(neto),
        },
    }


def get_adicionales_funebres(mes: str) -> List[Dict[str, Any]]:
    """Adicionales de Fúnebres.

    - Si existe definición exacta para el mes, se usa esa.
    - Si no existe, se toma la última definición anterior (prórroga automática).
      Esto permite, por ejemplo, que si el maestro quedó hasta 2026-01, en
      2026-02/03/04 se sigan ofreciendo los mismos adicionales.
    """
    idx = _build_index()
    mes_k = _mes_to_key(mes)

    d = idx.get("funebres_adic", {})
    if mes_k in d:
        return list(d.get(mes_k, []))

    # fallback: última definición <= mes_k
    keys = [k for k in d.keys() if isinstance(k, str) and k <= mes_k]
    if not keys:
        return []
    best = max(keys)
    return list(d.get(best, []))

def match_regla_conexiones(conexiones_o_nivel) -> Dict[str, Any]:
    """
    Agua Potable: reglas por umbrales (según tu UI):
    A: hasta 500
    B: 501-1000
    C: 1001-1600
    D: más de 1600
    El % es 7% encadenado (A=0%, B=7%, C=14,49%, D=22,5043%).
    """
    # Soporta dos entradas:
    # 1) cantidad (int) -> determina A/B/C/D por umbral
    # 2) nivel directo ("A"/"B"/"C"/"D") -> usa ese nivel
    level = 0
    cat = None
    label = None
    if isinstance(conexiones_o_nivel, str) and conexiones_o_nivel.strip():
        c = _norm(conexiones_o_nivel).upper()
        if c in ["A", "B", "C", "D"]:
            cat = c
            level = {"A": 0, "B": 1, "C": 2, "D": 3}[c]
            label = {
                "A": "A (hasta 500)",
                "B": "B (+7% s/A)",
                "C": "C (+7% s/B)",
                "D": "D (+7% s/C)",
            }[c]
        else:
            # Si viene un texto no esperado, intentamos tratarlo como número
            try:
                conexiones_o_nivel = int(c)
            except Exception:
                conexiones_o_nivel = 0

    if cat is None:
        try:
            n = int(conexiones_o_nivel)
        except Exception:
            n = 0
        if n <= 0:
            return {"cat": None, "pct": 0.0, "factor": 1.0, "label": None}

        if n <= 500:
            level = 0
            cat = "A"
            label = "A (hasta 500)"
        elif n <= 1000:
            level = 1
            cat = "B"
            label = "B (501 a 1000)"
        elif n <= 1600:
            level = 2
            cat = "C"
            label = "C (1001 a 1600)"
        else:
            level = 3
            cat = "D"
            label = "D (más de 1600)"

    factor = 1.07 ** level
    pct = factor - 1.0  # level 0 => 0
    return {"cat": cat, "pct": pct, "factor": factor, "label": label}

def get_titulo_pct_por_nivel(nivel: str) -> float:
    n = _norm(nivel).lower()
    if n in ["terciario", "terciario_turismo", "terciario (2.5%)", "2.5", "2,5"]:
        return 2.5
    if n in ["universitario", "licenciatura", "universitario (5%)", "5"]:
        return 5.0
    return 0.0

def get_regla_cajero(tipo: str) -> Dict[str, Any]:
    """
    Regla (CCT 130/75 - Acuerdo 26/09/1983):
      - Cajeros A y C: 12,25% sobre básico inicial Cajeros A
      - Cajeros B: 48% sobre básico inicial Cajeros B
    """
    t = _norm(tipo).upper()
    if t in ["A", "CAJERO A", "CAJEROS A", "CAJERO C", "CAJEROS C", "C"]:
        return {"tipo": t, "pct": 12.25}
    if t in ["B", "CAJERO B", "CAJEROS B"]:
        return {"tipo": t, "pct": 48.0}
    return {"tipo": t, "pct": 0.0}

def get_regla_km(categoria: str, km: float) -> Dict[str, Any]:
    """Normaliza el input de KM para el endpoint /regla-km.

    El motor (y/o el front) puede decidir si aplica la regla <=100 o >100.
    Acá devolvemos ambos tramos ya separados.
    """
    try:
        k = float(km or 0)
    except Exception:
        k = 0.0

    km_le_100 = min(k, 100.0)
    km_gt_100 = max(k - 100.0, 0.0)

    return {
        "categoria": _norm(categoria),
        "km": k,
        "km_le_100": km_le_100,
        "km_gt_100": km_gt_100,
    }


# ---------------------------
# Liquidación Final (Etapa 9)
# ---------------------------

def _parse_date_yyyy_mm_dd(s: str) -> _dt.date:
    """Parsea 'YYYY-MM-DD'. Lanza ValueError si es inválido."""
    if not s:
        raise ValueError('Fecha requerida')
    parts = s.strip().split('-')
    if len(parts) != 3:
        raise ValueError('Formato de fecha inválido (YYYY-MM-DD)')
    y, m, d = (int(parts[0]), int(parts[1]), int(parts[2]))
    return _dt.date(y, m, d)


def _years_complete(fecha_ing: _dt.date, fecha_egr: _dt.date) -> int:
    """Años completos entre fechas (antigüedad)."""
    y = fecha_egr.year - fecha_ing.year
    if (fecha_egr.month, fecha_egr.day) < (fecha_ing.month, fecha_ing.day):
        y -= 1
    return max(0, y)


def _months_diff(fecha_ing: _dt.date, fecha_egr: _dt.date) -> int:
    """Diferencia aproximada en meses completos (para reglas de preaviso / 245)."""
    m = (fecha_egr.year - fecha_ing.year) * 12 + (fecha_egr.month - fecha_ing.month)
    if fecha_egr.day < fecha_ing.day:
        m -= 1
    return max(0, m)


def _years_art245(fecha_ing: _dt.date, fecha_egr: _dt.date) -> int:
    """Años computables art. 245: 1 por año o fracción mayor a 3 meses."""
    total_months = _months_diff(fecha_ing, fecha_egr)
    if total_months < 3:
        return 0
    years = total_months // 12
    rem_months = total_months % 12
    if years == 0:
        return 1
    if rem_months > 3:
        years += 1
    return max(0, years)


def _vac_anuales_por_antig(anios: int) -> int:
    # LCT (estándar): <=5:14; >5<=10:21; >10<=20:28; >20:35
    if anios <= 5:
        return 14
    if anios <= 10:
        return 21
    if anios <= 20:
        return 28
    return 35


def calcular_final_payload(
    *,
    rama: str,
    agrup: str,
    categoria: str,
    jornada: float = 48.0,
    fecha_ingreso: str,
    fecha_egreso: str,
    tipo: str,
    mejor_rem: float = 0.0,
    mejor_nr: float = 0.0,
    mejor_total: float = 0.0,
    dias_mes: int = 0,
    vac_anuales: int = 0,
    vac_no_gozadas_dias: float = 0.0,
    preaviso_dias: int = 0,
    integracion: bool = True,
    sac_sobre_preaviso: bool = False,
    sac_sobre_integracion: bool = True,
    osecac: bool = True,
    afiliado: bool = False,
    sind_pct: float = 0.0,
    sind_fijo: float = 0.0,
    # Extras del mes de baja (mismos parámetros que mensual)
    titulo_pct: float = 0.0,
    zona_pct: float = 0.0,
    fer_no_trab: int = 0,
    fer_trab: int = 0,
    vac_goz: int = 0,
    aus_inj: int = 0,
    susp_dias: int = 0,
    hex50: float = 0.0,
    hex100: float = 0.0,
    hs_noct: float = 0.0,
    km_tipo: str = "",
    km_menos100: int = 0,
    km_mas100: int = 0,
    a_cuenta_rem: float = 0.0,
    viaticos_nr: float = 0.0,
    manejo_caja: bool = False,
    cajero_tipo: str = "",
    faltante_caja: float = 0.0,
    armado_vidriera: bool = False,
    adelanto_sueldo: float = 0.0,
    fun_adic: str = "",
    jubilado: bool = False,
    embargo: float = 0.0,
) -> Dict[str, Any]:
    """Liquidación final básica (MVP Etapa 9).

    - Usa MRMNH / Mejor salario como base (Rem + NR) y prorratea por día.
    - SAC proporcional por días de semestre (base * días / 360).
    - Vacaciones no gozadas (días) editable, pago 1/25, como indemnizatorio + SAC s/vac (sin descuentos).
    - Despido sin causa: incluye indemnización art. 245 + preaviso (opción) + integración (opción).

    Devuelve items con columnas: r (rem), n (no rem), i (indemnizatorio), d (descuentos).
    """
    import calendar as _cal

    fi = _parse_date_yyyy_mm_dd(fecha_ingreso)
    fe = _parse_date_yyyy_mm_dd(fecha_egreso)
    if fe < fi:
        raise ValueError('La fecha de egreso no puede ser anterior a la de ingreso')

    # Base mejor salario
    mr = float(mejor_rem or 0.0)
    nn = float(mejor_nr or 0.0)
    mt = float(mejor_total or 0.0)

    if mt <= 0:
        mt = mr + nn
    if mr <= 0 and nn <= 0 and mt > 0:
        # si solo vino total, lo tratamos como Rem (compatibilidad)
        mr = mt
        nn = 0.0

    base_total = max(0.0, mr + nn)
    if base_total <= 0:
        raise ValueError('MRMNH / Mejor salario debe ser mayor a 0')

    share_rem = mr / base_total if base_total else 1.0
    share_nr = nn / base_total if base_total else 0.0

    # Días del mes de egreso
    dim = _cal.monthrange(fe.year, fe.month)[1]
    dia_baja = fe.day

    # Días trabajados del mes (criterio para evitar "inflar" o "achicar" meses de 28/29/31):
    # - Si el trabajador estuvo TODO el mes (desde el 1 hasta el último día del mes de baja),
    #   se liquida como mes completo (30/30), independientemente de que el mes tenga 28, 29, 30 o 31 días.
    # - Si NO es mes completo, se prorratea por días reales trabajados en ese mes, siempre sobre divisor 30.
    #   (Práctica estándar para mensualizados: divisor 30).
    mes_inicio = _dt.date(fe.year, fe.month, 1)
    mes_fin = _dt.date(fe.year, fe.month, dim)
    inicio_trab_mes = max(fi, mes_inicio)
    dm_real = (fe - inicio_trab_mes).days + 1 if fe >= inicio_trab_mes else 0
    dm_real = max(0, min(dim, int(dm_real)))
    mes_completo = (inicio_trab_mes == mes_inicio) and (fe == mes_fin)
    dm_prorr = 30 if mes_completo else dm_real

    # Antigüedad
    anios_antig = _years_complete(fi, fe)
    anios_245 = _years_art245(fi, fe)

    # Vacaciones
    vac_an = int(vac_anuales or 0)
    if vac_an <= 0:
        vac_an = _vac_anuales_por_antig(anios_antig)

    # Días trabajados del año (para sugerir vacaciones)
    start_year = max(fi, _dt.date(fe.year, 1, 1))
    dias_anio = (fe - start_year).days + 1
    dias_anio = max(0, dias_anio)

    # Sugerencia de vacaciones por proporción, pero el usuario define los días finales
    vac_sugeridas = round2(vac_an * (dias_anio / 365.0)) if dias_anio else 0.0
    vac_no_goz = float(vac_no_gozadas_dias or 0.0)
    if vac_no_goz <= 0:
        vac_no_goz = float(vac_sugeridas or 0.0)

    # Días trabajados del semestre (para SAC proporcional)
    sem_start = _dt.date(fe.year, 1, 1) if fe.month <= 6 else _dt.date(fe.year, 7, 1)
    start_sem = max(fi, sem_start)
    dias_sem = (fe - start_sem).days + 1
    dias_sem = max(0, dias_sem)

    # Helpers de prorrateo
    def split_amount(total: float) -> Tuple[float, float]:
        t = round2(total)
        r = round2(t * share_rem)
        n = round2(t - r)
        return r, n

    def _antig_pct_rama(_rama: str, _anios: int) -> float:
        """Porcentaje de antigüedad según rama.

        - Agua Potable: 2% anual acumulativo
        - Resto: 1% por año (no acumulativo)
        """
        try:
            r0 = norm_rama(_rama)
        except Exception:
            r0 = str(_rama or '').strip().upper()

        a = max(0, int(_anios or 0))
        if r0 in ("AGUA POTABLE", "AGUA", "AGUAPOTABLE"):
            # 2% anual acumulativo
            return (pow(1.02, a) - 1.0) if a else 0.0
        return 0.01 * float(a)

    def _desglosar_base(total: float, pct_ant: float) -> Tuple[float, float, float]:
        """Desglosa un total de (Básico + Antigüedad + Presentismo) en sus 3 componentes.

        Fórmula usada en mensual:
        - Antig = Básico * pct
        - Pres = (Básico + Antig) / 12
        - Total = Básico + Antig + Pres

        Ajusta el residuo por redondeo en Presentismo para que la suma coincida.
        """
        t = round2(float(total or 0.0))
        if t <= 0:
            return 0.0, 0.0, 0.0
        pct = max(0.0, float(pct_ant or 0.0))
        denom = (1.0 + pct) * (13.0 / 12.0)
        if denom <= 0:
            return t, 0.0, 0.0
        bas = round2(t / denom)
        ant = round2(bas * pct) if pct else 0.0
        pres = round2((bas + ant) / 12.0) if (bas or ant) else 0.0
        # Ajuste por redondeo: todo el residuo a Presentismo
        resid = round2(t - round2(bas + ant + pres))
        if resid:
            pres = round2(pres + resid)
        return bas, ant, pres

    # -------- Mes de baja (Liquidación del mes) --------
    # El MEJOR SALARIO (base indemnizatoria) no pierde presentismo.
    # En la liquidación del mes (y su integración), si hubo 2+ ausencias injustificadas,
    # se pierde el presentismo de ese mes.
    pct_ant_final = _antig_pct_rama(rama, anios_antig)
    bas_full_r, ant_full_r, pres_full_r = _desglosar_base(mr, pct_ant_final)
    bas_full_n, ant_full_n, pres_full_n = _desglosar_base(nn, pct_ant_final)

    aus_dias = max(0, int(aus_inj or 0))
    presentismo_habil_baja = aus_dias < 2

    base_mes_r = mr if presentismo_habil_baja else round2(mr - pres_full_r)
    base_mes_n = nn if presentismo_habil_baja else round2(nn - pres_full_n)
    base_mes_r = max(0.0, base_mes_r)
    base_mes_n = max(0.0, base_mes_n)

    frac_mes = (float(dm_prorr) / 30.0) if dm_prorr else 0.0
    hab_mes_r = round2(base_mes_r * frac_mes)
    hab_mes_n = round2(base_mes_n * frac_mes)
    hab_mes_total = round2(hab_mes_r + hab_mes_n)

    # Vacaciones no gozadas (indemnizatorio): pago 1/25 sobre base total (incluye NR)
    vac_pago_total = round2((base_total / 25.0) * vac_no_goz) if vac_no_goz else 0.0

    # SAC sobre vacaciones no gozadas (indemnizatorio)
    sac_vac_total = round2(vac_pago_total / 12.0) if vac_pago_total else 0.0
    # SAC proporcional del semestre
    sac_prop_total = round2(base_total * (dias_sem / 360.0)) if dias_sem else 0.0
    sac_prop_r, sac_prop_n = split_amount(sac_prop_total)

    # Integración mes despido (art. 233) – default ON
    dias_int = max(0, dim - dia_baja) if integracion else 0
    # Integración del mes (mismo criterio de presentismo que el mes de baja)
    frac_int = (dias_int / 30.0) if dias_int else 0.0
    integ_r = round2(base_mes_r * frac_int) if dias_int else 0.0
    integ_n = round2(base_mes_n * frac_int) if dias_int else 0.0
    integ_total = round2(integ_r + integ_n)

    sac_integ_total = round2(integ_total / 12.0) if (sac_sobre_integracion and integ_total) else 0.0
    if sac_integ_total and (integ_r + integ_n) > 0:
        sac_integ_r = round2(sac_integ_total * (integ_r / (integ_r + integ_n)))
        sac_integ_n = round2(sac_integ_total - sac_integ_r)
    else:
        sac_integ_r, sac_integ_n = 0.0, 0.0

    # Preaviso (art. 231/232) – indemnizatorio
    prev_dias = int(preaviso_dias or 0)
    prev_total = round2(base_total * (prev_dias / 30.0)) if prev_dias else 0.0
    sac_prev_total = round2(prev_total / 12.0) if (sac_sobre_preaviso and prev_total) else 0.0

    # Indemnización antigüedad (art. 245)
    ind_antig = 0.0
    # Indemnización por fallecimiento (art. 248)
    ind_fall = 0.0
    tipo_n = (tipo or '').strip().upper()
    if tipo_n in ('DESPIDO_SIN_CAUSA', 'DESPIDO SIN CAUSA', 'SIN_CAUSA'):
        ind_antig = round2(base_total * float(anios_245 or 0))

    if tipo_n == 'FALLECIMIENTO':
        # Art. 248: indemnización por fallecimiento = 50% de la indemnización art. 245
        ind_fall = round2(base_total * float(anios_245 or 0) * 0.5)

    # Armado de items
    def item(concepto: str, r: float = 0.0, n: float = 0.0, i: float = 0.0, d: float = 0.0, base_num: float = 0.0) -> Dict[str, Any]:
        out = {"concepto": concepto, "r": float(r), "n": float(n), "i": float(i), "d": float(d)}
        if base_num:
            out["base"] = float(base_num)
        return out

    items: List[Dict[str, Any]] = []

    # Desglose de DÍAS TRABAJADOS (Básico / Antigüedad / Presentismo), manteniendo totales.
    pct_ant_final = _antig_pct_rama(rama, anios_antig)
    bas_full_r, ant_full_r, pres_full_r = _desglosar_base(mr, pct_ant_final)
    bas_full_n, ant_full_n, pres_full_n = _desglosar_base(nn, pct_ant_final)

    def _prorratear_componentes(
        dias: int,
        total_r_obj: float,
        total_n_obj: float,
        label_base: str,
        ctx: str,
        base_num_first: float,
        incluir_presentismo: bool = True,
    ) -> None:
        """Agrega filas de desglose.

        - Si incluir_presentismo=True: 3 filas (Básico / Antigüedad / Presentismo).
        - Si incluir_presentismo=False: 2 filas (Básico / Antigüedad) y NO se muestra Presentismo.

        Se prorratea por días (sobre 30) y se ajusta el residuo por redondeo para que
        la suma coincida EXACTO con los totales objetivo (Rem/NR).
        """
        if dias <= 0:
            return
        frac = float(dias) / 30.0

        b_r = round2(bas_full_r * frac)
        a_r = round2(ant_full_r * frac)
        b_n = round2(bas_full_n * frac)
        a_n = round2(ant_full_n * frac)

        if incluir_presentismo:
            p_r = round2(pres_full_r * frac)
            p_n = round2(pres_full_n * frac)

            # Ajustes por redondeo (residuo va a Presentismo)
            dr = round2(round2(total_r_obj) - round2(b_r + a_r + p_r))
            if dr:
                p_r = round2(p_r + dr)
            dn = round2(round2(total_n_obj) - round2(b_n + a_n + p_n))
            if dn:
                p_n = round2(p_n + dn)

            # Evitar presentismo negativo por ajustes de redondeo
            if p_r < 0:
                b_r = round2(b_r + p_r)
                p_r = 0.0
            if p_n < 0:
                b_n = round2(b_n + p_n)
                p_n = 0.0

            items.append(item(label_base, r=b_r, n=b_n, base_num=base_num_first))
            # Base para auditoría:
            # - Antigüedad: se calcula sobre Básico (prorrateado)
            # - Presentismo: se calcula sobre (Básico + Antigüedad) (prorrateados)
            base_ant = round2(b_r + b_n)
            base_pre = round2(b_r + b_n + a_r + a_n)
            items.append(item(f"Antigüedad ({ctx})", r=a_r, n=a_n, base_num=base_ant))
            items.append(item(f"Presentismo ({ctx})", r=p_r, n=p_n, base_num=base_pre))
            return

        # Sin presentismo: ajuste por redondeo contra (Básico + Antigüedad)
        dr2 = round2(round2(total_r_obj) - round2(b_r + a_r))
        if dr2:
            a_r = round2(a_r + dr2)
        dn2 = round2(round2(total_n_obj) - round2(b_n + a_n))
        if dn2:
            a_n = round2(a_n + dn2)

        # Evitar negativos por ajustes
        if a_r < 0:
            b_r = round2(max(0.0, b_r + a_r))
            a_r = 0.0
        if a_n < 0:
            b_n = round2(max(0.0, b_n + a_n))
            a_n = 0.0

        items.append(item(label_base, r=b_r, n=b_n, base_num=base_num_first))
        # Sin presentismo, la antigüedad sigue basándose en el Básico prorrateado.
        items.append(item(f"Antigüedad ({ctx})", r=a_r, n=a_n, base_num=round2(b_r + b_n)))

    if mes_completo:
        suf_dm = f"mes completo ({dim} días)"
    else:
        suf_dm = f"{dm_real} día{'s' if dm_real != 1 else ''}"

    _prorratear_componentes(
        dm_prorr,
        hab_mes_r,
        hab_mes_n,
        f"Días trabajados del mes ({suf_dm}) — Básico",
        f"Días trabajados del mes ({suf_dm})",
        base_total,
        incluir_presentismo=presentismo_habil_baja,
    )

    if vac_pago_total:
        items.append(item(f"Vacaciones no gozadas (Indem.)", i=vac_pago_total, base_num=base_total))
        if sac_vac_total:
            items.append(item("SAC s/ Vacaciones no gozadas (Indem.)", i=sac_vac_total, base_num=vac_pago_total))

    if sac_prop_total:
        items.append(item("SAC proporcional", r=sac_prop_r, n=sac_prop_n, base_num=base_total))

    if integ_total:
        suf_int = f"{dias_int} día{'s' if dias_int != 1 else ''}"
        _prorratear_componentes(
            dias_int,
            integ_r,
            integ_n,
            f"Integración mes despido ({suf_int}) — Básico",
            f"Integración mes despido ({suf_int})",
            base_total,
            incluir_presentismo=presentismo_habil_baja,
        )
        if sac_integ_total:
            items.append(item("SAC s/ integración", r=sac_integ_r, n=sac_integ_n, base_num=integ_total))


    if prev_total:
        items.append(item(f"Preaviso ({prev_dias} día{'s' if prev_dias != 1 else ''})", i=prev_total, base_num=base_total))
        if sac_prev_total:
            items.append(item("SAC s/ preaviso", i=sac_prev_total, base_num=prev_total))

    if ind_fall:
        items.append(item("Indemnización por fallecimiento (Art. 248)", i=ind_fall, base_num=base_total))

    if ind_antig:
        items.append(item(f"Indemnización antigüedad (Art. 245) ({anios_245} año{'s' if anios_245 != 1 else ''})", i=ind_antig, base_num=base_total))

    # -----------------
    # Extras del mes de baja (mismos conceptos que en mensual, pero dentro de Liquidación Final)
    # Nota: se calculan reutilizando el motor mensual del mes de egreso, y se agregan SOLO los
    # conceptos distintos de Básico/Antigüedad/Presentismo y deducciones.
    # -----------------
    try:
        mes_baja = f"{fe.year:04d}-{fe.month:02d}"

        # Fúnebres: lista de adicionales seleccionados (separados por ';')
        fun_list: List[str] = []
        if fun_adic:
            fun_list = [x.strip() for x in str(fun_adic).replace(",", ";").split(";") if x.strip()]

        mensual = calcular_payload(
            rama=rama,
            agrup=agrup,
            categoria=categoria,
            mes=mes_baja,
            jornada=jornada,
            anios_antig=anios_antig,
            osecac=osecac,
            afiliado=afiliado,
            sind_pct=float(sind_pct or 0),
            titulo_pct=float(titulo_pct or 0),
            zona_pct=float(zona_pct or 0),
            fer_no_trab=int(fer_no_trab or 0),
            fer_trab=int(fer_trab or 0),
			vac_goz=int(vac_goz or 0),
            aus_inj=int(aus_inj or 0),
            susp_dias=int(susp_dias or 0),
            hex50=float(hex50 or 0),
            hex100=float(hex100 or 0),
            hs_noct=float(hs_noct or 0),
            km_tipo=str(km_tipo or ""),
            km_menos100=int(km_menos100 or 0),
            km_mas100=int(km_mas100 or 0),
            a_cuenta_rem=float(a_cuenta_rem or 0),
            viaticos_nr=float(viaticos_nr or 0),
            manejo_caja=bool(manejo_caja),
            cajero_tipo=str(cajero_tipo or ""),
            faltante_caja=float(faltante_caja or 0),
            armado_vidriera=bool(armado_vidriera),
            adelanto_sueldo=float(adelanto_sueldo or 0),
            jubilado=bool(jubilado),
            embargo=0.0,
            fun_adic=(";".join(fun_list) if fun_list else ""),
        )

        def _skip_concepto(con: str) -> bool:
            c = (con or "").strip().lower()
            if not c:
                return True
            # excluir bases y deducciones (ya se calculan en final)
            claves = [
                "básico", "basico", "antigüedad", "antiguedad", "presentismo",
                "incr.", "recomp.",
                "jubil", "pami", "obra social", "osecac", "faecys", "sindicato",
                "embargo", "total", "neto",
            ]
            return any((c == k) or c.startswith(k) or (k in c and k not in ("incr.", "recomp.")) for k in claves)

        for it in (mensual or {}).get("items", []) or []:
            con = str(it.get("concepto", ""))
            if _skip_concepto(con):
                continue
            r = float(it.get("r", 0.0) or 0.0)
            n = float(it.get("n", 0.0) or 0.0)
            d = float(it.get("d", 0.0) or 0.0)
            if abs(r) + abs(n) + abs(d) <= 0:
                continue
            # Preservar base numérica (ej. valor hora en horas extra) si viene del mensual
            base_num = 0.0
            try:
                if it.get("base") not in (None, ""):
                    base_num = float(it.get("base"))
            except Exception:
                base_num = 0.0
            items.append(item(con, r=r, n=n, d=d, base_num=base_num))
    except Exception:
        # Si algo falla, no frenamos la liquidación final.
        pass

    # Totales antes de descuentos
    rem_total = round2(sum(x.get('r', 0.0) for x in items))
    nr_total = round2(sum(x.get('n', 0.0) for x in items))
    ind_total = round2(sum(x.get('i', 0.0) for x in items))

    # -----------------
    # Deducciones (misma lógica que mensual)
    # -----------------
    rem_aportes = rem_total

    os_base = rem_aportes  # base mostrada para Obra Social

    jub = 0.0
    pami = 0.0
    os_aporte = 0.0
    osecac_100 = 0.0
    sind = 0.0
    sind_fijo_monto = 0.0

    if bool(jubilado):
        # Jubilado: según criterio del sistema vigente
        jub = round2(rem_aportes * 0.11)
        base_fs = round2(rem_aportes + nr_total)
        faecys = round2(base_fs * 0.005)
        sind_solid = round2(base_fs * 0.02)
        if bool(afiliado):
            if float(sind_pct or 0) > 0:
                sind = round2(base_fs * (float(sind_pct) / 100.0))
            if float(sind_fijo or 0) > 0:
                sind_fijo_monto = round2(float(sind_fijo))
    else:
        jub = round2(rem_aportes * 0.11)
        pami = round2(rem_aportes * 0.03)

        # Obra Social: base jornada completa (48hs) para TODAS las ramas (sin prorrateo por jornada).
        try:
            j_in = float(jornada or 48.0)
        except Exception:
            j_in = 48.0
        if j_in <= 0:
            j_in = 48.0
        # Obra Social: base jornada completa (48hs) para TODAS las ramas (sin prorrateo por jornada).
        factor_os = 1.0 if j_in >= 48.0 else (48.0 / j_in)
        rem_aportes_os = round2(rem_aportes * factor_os)
        nr_os = round2(nr_total * factor_os)
        os_base = round2((rem_aportes_os + nr_os) if bool(osecac) else rem_aportes_os)
        os_aporte = round2(os_base * 0.03) if bool(osecac) else 0.0
        osecac_100 = 100.0 if bool(osecac) else 0.0

        if bool(afiliado):
            if float(sind_pct or 0) > 0:
                sind = round2((rem_aportes + nr_total) * (float(sind_pct) / 100.0))
            if float(sind_fijo or 0) > 0:
                sind_fijo_monto = round2(float(sind_fijo))

    # Total de deducciones (para neto): debe contemplar todos los conceptos
    # que efectivamente se agregan a la columna de Deducciones.
    if bool(jubilado):
        ded_pre = round2(jub + faecys + sind_solid + sind + sind_fijo_monto)
    else:
        ded_pre = round2(jub + pami + os_aporte + osecac_100 + sind + sind_fijo_monto)

    neto_pre = round2((rem_total + nr_total + ind_total) - ded_pre)

    emb_in = 0.0
    try:
        emb_in = float(embargo or 0.0)
    except Exception:
        emb_in = 0.0
    emb_in = max(0.0, emb_in)
    embargo_monto = round2(min(emb_in, max(0.0, neto_pre))) if emb_in else 0.0

    ded_total = round2(ded_pre + embargo_monto)
    neto = round2(neto_pre - embargo_monto)

    # Agregar filas de descuentos al final
    if bool(jubilado):
        base_fs = round2(rem_aportes + nr_total)
        faecys = round2(base_fs * 0.005)
        sind_solid = round2(base_fs * 0.02)
        if jub:
            items.append(item("Jubilación 11% (Jubilado)", d=jub, base_num=rem_aportes))
        if faecys:
            items.append(item("FAECYS 0,5%", d=faecys, base_num=base_fs))
        if sind_solid:
            items.append(item("Sindicato 2%", d=sind_solid, base_num=base_fs))
        if sind:
            items.append(item(f"Afiliación {float(sind_pct):g}%", d=sind, base_num=base_fs))
        if sind_fijo_monto:
            items.append(item("Afiliación (fijo)", d=sind_fijo_monto))
    else:
        if jub:
            items.append(item("Jubilación 11%", d=jub, base_num=rem_aportes))
        if pami:
            items.append(item("Ley 19.032 (PAMI) 3%", d=pami, base_num=rem_aportes))
        items.append(item("Obra Social 3%", d=os_aporte, base_num=os_base))
        if osecac_100:
            items.append(item("OSECAC $100", d=osecac_100))
        if sind:
            items.append(item(f"Sindicato {float(sind_pct):g}%", d=sind, base_num=(rem_aportes + nr_total)))
        if sind_fijo_monto:
            items.append(item("Sindicato (fijo)", d=sind_fijo_monto))

    if embargo_monto:
        items.append(item("Embargo (desc.)", d=embargo_monto, base_num=neto_pre))

    # Recalcular totales para que incluyan filas de descuentos en el cuerpo
    rem_total = round2(sum(x.get('r', 0.0) for x in items))
    nr_total = round2(sum(x.get('n', 0.0) for x in items))
    ind_total = round2(sum(x.get('i', 0.0) for x in items))

    return {
        "ok": True,
        "rama": _norm(rama),
        "agrup": _norm(agrup),
        "categoria": _norm(categoria),
        "tipo": tipo,
        "fecha_ingreso": fecha_ingreso,
        "fecha_egreso": fecha_egreso,
        "anios_antig": anios_antig,
        "anios_245": anios_245,
        # Días calendario trabajados dentro del mes de egreso
        "dias_mes": dm_real,
        # Días usados para prorrateo (sobre divisor 30). Si es mes completo, vale 30.
        "dias_prorr": dm_prorr,
        "mes_completo": bool(mes_completo),
        "dias_semestre": dias_sem,
        "vac_anuales": vac_an,
        "vac_no_gozadas_dias": vac_no_goz,
        "vac_sugeridas": vac_sugeridas,
        "items": items,
        "totales": {
            "rem": rem_total,
            "nr": nr_total,
            "ind": ind_total,
            "ded": ded_total,
            "neto": neto,
        },
    }
