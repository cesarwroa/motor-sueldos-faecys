import base64
from datetime import datetime, timezone
import hashlib
import hmac
from io import BytesIO
import json
import mimetypes
import os
import re
import time
import unicodedata
import uuid

from fastapi import FastAPI, File, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path

from pydantic import BaseModel
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from escalas import (
    get_meta,
    get_payload,
    calcular_payload,
    get_adicionales_funebres,
    match_regla_conexiones,
    get_titulo_pct_por_nivel,
    get_regla_cajero,
    get_regla_km,
    calcular_final_payload,
    calcular_vacaciones_payload,
)

app = FastAPI(
    title="motor-sueldos-faecys",
    docs_url=None,
    redoc_url=None,
    openapi_url=None,
)

# CORS
ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "ALLOWED_ORIGINS",
        "https://app.calculadoradecomercio.com.ar,"
        "https://calculadoradecomercio.com.ar,"
        "https://www.calculadoradecomercio.com.ar",
    ).split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)

BASE_DIR = Path(__file__).resolve().parent
ADMIN_LOGIN_EMAIL = os.getenv("ADMIN_LOGIN_EMAIL", "").strip().lower()
ADMIN_LOGIN_PASSWORD = os.getenv("ADMIN_LOGIN_PASSWORD", "")
ADMIN_ACCESS_SECRET = os.getenv("ADMIN_ACCESS_SECRET", "")
ADMIN_TOKEN_TTL_SECONDS = int(os.getenv("ADMIN_TOKEN_TTL_SECONDS", "43200"))
ADMIN_FEATURES_FILE = BASE_DIR / "data" / "admin_features.json"
ADMIN_COMPANIES_FILE = BASE_DIR / "data" / "admin_companies.json"
ADMIN_COMPANY_STATE_FILE = BASE_DIR / "data" / "admin_company_state.json"
ADMIN_EMPLOYEES_FILE = BASE_DIR / "data" / "admin_employees.json"
ADMIN_EMPLOYEE_STATE_FILE = BASE_DIR / "data" / "admin_employee_state.json"
ADMIN_LEADS_FILE = BASE_DIR / "data" / "admin_leads.json"
ADMIN_PAYROLL_HISTORY_FILE = BASE_DIR / "data" / "admin_payroll_history.json"
ADMIN_ARCA_MAPPINGS_FILE = BASE_DIR / "data" / "admin_arca_mappings.json"
ADMIN_COMPANY_ASSETS_DIR = BASE_DIR / "data" / "admin_company_assets"
ADMIN_COMPANY_ASSET_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
ADMIN_COMPANY_ASSET_MAX_BYTES = int(os.getenv("ADMIN_COMPANY_ASSET_MAX_BYTES", str(5 * 1024 * 1024)))
PUBLIC_INDEX_FILE = BASE_DIR / "public_index.html"
PUBLIC_STATIC_INDEX_FILE = BASE_DIR / "public" / "index.html"
COMPANY_PORTAL_FILE = BASE_DIR / "public" / "empresas.html"
EMPLOYEE_IMPORT_TEMPLATE_FILE = BASE_DIR / "public" / "plantilla_importacion_empleados.xlsx"
COMPANY_PAYROLL_MANUAL_FILE = BASE_DIR / "public" / "manual-liquidacion-recibos-libro-sueldos-digital-arca.pdf"
ADMIN_INDEX_FILE = BASE_DIR / "index.html"
NOINDEX_HEADERS = {"X-Robots-Tag": "noindex, nofollow, noarchive"}
FEATURE_ACCESS_ALLOWED = {"off", "admin_only", "public"}
FEATURE_PUBLIC_MAP = {
    "liquidacion_final": "liquidacion_final_publica",
}
DEFAULT_FEATURE_ACCESS = {
    "liquidacion_final": "public",
    "registro_empresas": "admin_only",
    "anexo_costo_empleador": "public",
    "portal_empresa": "admin_only",
    "firma_digital": "admin_only",
    "portal_empleado": "admin_only",
    "gestion_nomina": "admin_only",
}
DEFAULT_PUBLIC_FEATURES = {
    "liquidacion_final_publica": True,
}

EMPLOYEE_IMPORT_MAX_BYTES = 3 * 1024 * 1024
EMPLOYEE_IMPORT_MAX_ROWS = 500
_RATE_LIMIT_BUCKETS: Dict[str, List[float]] = {}


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; base-uri 'self'; frame-ancestors 'none'; "
        "object-src 'none'; form-action 'self'; "
        "img-src 'self' data: blob: "
        "https://*.google-analytics.com https://*.googletagmanager.com; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline' https://www.googletagmanager.com; "
        "connect-src 'self' https://calculadoradecomercio.com.ar "
        "https://*.google-analytics.com https://*.analytics.google.com "
        "https://*.googletagmanager.com"
    )
    return response


def _client_ip(request: Request) -> str:
    forwarded = str(request.headers.get("x-forwarded-for") or "").split(",", 1)[0].strip()
    return forwarded or (request.client.host if request.client else "unknown")


def _enforce_rate_limit(request: Request, scope: str, limit: int, window_seconds: int) -> None:
    now = time.monotonic()
    key = f"{scope}:{_client_ip(request)}"
    cutoff = now - window_seconds
    attempts = [stamp for stamp in _RATE_LIMIT_BUCKETS.get(key, []) if stamp > cutoff]
    if len(attempts) >= limit:
        raise HTTPException(status_code=429, detail="Demasiadas solicitudes. Intentá nuevamente más tarde.")
    attempts.append(now)
    _RATE_LIMIT_BUCKETS[key] = attempts


def _require_admin_security_config() -> None:
    if (
        not ADMIN_LOGIN_EMAIL
        or len(ADMIN_LOGIN_PASSWORD) < 12
        or len(ADMIN_ACCESS_SECRET.encode("utf-8")) < 32
    ):
        raise HTTPException(
            status_code=503,
            detail="El acceso administrativo no está configurado de forma segura.",
        )
EMPLOYEE_IMPORT_HEADERS = {
    "legajo": "file_number",
    "apellido_y_nombre": "full_name",
    "cuil": "cuil",
    "dni": "dni",
    "fecha_nacimiento": "birth_date",
    "fecha_ingreso": "start_date",
    "fecha_egreso": "end_date",
    "estado": "status",
    "rama": "rama",
    "agrupamiento": "agrup",
    "categoria": "category",
    "jornada": "workday",
    "horas_semanales": "weekly_hours",
    "modalidad_contractual": "contract_modality",
    "codigo_modalidad_arca": "contract_modality_code",
    "email": "email",
    "telefono": "phone",
    "domicilio": "address",
    "localidad": "locality",
    "provincia": "province",
    "obra_social": "health_insurance_name",
    "codigo_obra_social_arca": "health_insurance_code",
    "numero_afiliado_obra_social": "health_insurance_member",
    "vigencia_obra_social_desde": "health_insurance_start",
    "sindicato": "union_name",
    "codigo_sindical": "union_code",
    "numero_afiliado_sindical": "union_member",
    "banco": "bank_name",
    "cbu": "cbu",
    "alias_bancario": "bank_alias",
    "forma_pago": "payment_method",
}


class AdminLoginRequest(BaseModel):
    email: str
    password: str


class AdminFeaturesUpdate(BaseModel):
    liquidacion_final: Optional[str] = None
    registro_empresas: Optional[str] = None
    anexo_costo_empleador: Optional[str] = None
    portal_empresa: Optional[str] = None
    firma_digital: Optional[str] = None
    portal_empleado: Optional[str] = None
    gestion_nomina: Optional[str] = None
    liquidacion_final_publica: Optional[bool] = None


class AdminCompanyCreate(BaseModel):
    razon_social: str
    cuit: str = ""
    rama: str = ""
    email: str = ""
    telefono: str = ""
    actividad: str = ""
    logo_url: str = ""
    direccion_calle: str = ""
    direccion_numero: str = ""
    direccion_piso: str = ""
    direccion_depto: str = ""
    localidad: str = ""
    provincia: str = ""
    codigo_postal: str = ""
    estado: str = "prueba"
    observaciones: str = ""


class AdminCompanyActiveUpdate(BaseModel):
    company_id: str = ""


class AdminEmployeeCreate(BaseModel):
    company_id: str
    legajo: str = ""
    apellido_nombre: str
    cuil: str = ""
    sueldo_jornal: str = ""
    categoria: str = ""
    tarea: str = ""
    fecha_ingreso: str = ""
    obra_social: str = ""
    obra_social_periodo: str = ""
    deposito_previsional: str = ""
    deposito_previsional_fecha: str = ""
    lugar_pago: str = ""
    estado: str = "prueba"
    observaciones: str = ""


class AdminEmployeeActiveUpdate(BaseModel):
    company_id: str
    employee_id: str = ""


class AdminPayrollHistoryCreate(BaseModel):
    company_id: str
    employee_id: str
    periodo: str
    tipo: str = "mensual"
    inputs: Dict[str, Any] = {}
    resultado: Dict[str, Any]


class AdminArcaMappingCreate(BaseModel):
    company_id: str
    concepto: str
    codigo_empleador: str
    codigo_arca: str
    unidad: str = "$"


class PublicLeadCreate(BaseModel):
    nombre: str
    email: str
    empresa: str
    telefono: str = ""
    empleados: str = ""
    motivo: str = "Consulta empresa / nomina"
    mensaje: str = ""


def _sanitize_admin_asset_stem(value: str) -> str:
    stem = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(value or "").strip().lower())
    return stem.strip("-_") or "logo"


def _build_admin_company_asset_url(asset_name: str) -> str:
    return f"/admin/company-assets/{asset_name}"


def _b64url_encode(raw: bytes) -> str:
    return base64.urlsafe_b64encode(raw).rstrip(b"=").decode("ascii")


def _b64url_decode(raw: str) -> bytes:
    padding = "=" * (-len(raw) % 4)
    return base64.urlsafe_b64decode((raw + padding).encode("ascii"))


def _sign_admin_token(payload: Dict[str, Any]) -> str:
    payload_raw = json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
    payload_b64 = _b64url_encode(payload_raw)
    signature = hmac.new(
        ADMIN_ACCESS_SECRET.encode("utf-8"),
        payload_b64.encode("ascii"),
        hashlib.sha256,
    ).digest()
    return f"{payload_b64}.{_b64url_encode(signature)}"


def _issue_admin_token(email: str) -> str:
    now = int(time.time())
    payload = {
        "email": email,
        "role": "admin",
        "iat": now,
        "exp": now + ADMIN_TOKEN_TTL_SECONDS,
    }
    return _sign_admin_token(payload)


def _read_admin_token(token: str) -> Dict[str, Any]:
    try:
        payload_b64, signature_b64 = token.split(".", 1)
    except ValueError as exc:
        raise HTTPException(status_code=401, detail="Token admin inválido.") from exc

    expected_signature = hmac.new(
        ADMIN_ACCESS_SECRET.encode("utf-8"),
        payload_b64.encode("ascii"),
        hashlib.sha256,
    ).digest()
    actual_signature = _b64url_decode(signature_b64)

    if not hmac.compare_digest(expected_signature, actual_signature):
        raise HTTPException(status_code=401, detail="Firma de sesión inválida.")

    try:
        payload = json.loads(_b64url_decode(payload_b64).decode("utf-8"))
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=401, detail="No se pudo leer la sesión admin.") from exc

    exp = int(payload.get("exp") or 0)
    if exp <= int(time.time()):
        raise HTTPException(status_code=401, detail="La sesión admin venció.")

    if str(payload.get("role") or "").lower() != "admin":
        raise HTTPException(status_code=401, detail="La sesión no tiene permisos de administrador.")

    return payload


def _extract_admin_token(authorization: Optional[str]) -> str:
    if not authorization:
        raise HTTPException(status_code=401, detail="Falta el token admin.")
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token.strip():
        raise HTTPException(status_code=401, detail="Formato de autorización inválido.")
    return token.strip()


def _default_feature_store() -> Dict[str, Any]:
    feature_access = dict(DEFAULT_FEATURE_ACCESS)
    return {
        "feature_access": feature_access,
        "public_features": _feature_access_to_public_features(feature_access),
        "updated_at": None,
        "updated_by": "",
    }


def _normalize_feature_access_value(value: Any, default: str) -> str:
    raw = str(value or "").strip().lower()
    if raw in FEATURE_ACCESS_ALLOWED:
        return raw
    return default


def _feature_access_to_public_features(feature_access: Dict[str, str]) -> Dict[str, bool]:
    public_features = dict(DEFAULT_PUBLIC_FEATURES)
    for feature_name, public_key in FEATURE_PUBLIC_MAP.items():
        public_features[public_key] = str(feature_access.get(feature_name) or "").strip().lower() == "public"
    return public_features


def _normalize_feature_store(raw: Any) -> Dict[str, Any]:
    store = _default_feature_store()
    if not isinstance(raw, dict):
        return store

    raw_access = raw.get("feature_access")
    if isinstance(raw_access, dict):
        for key, default_value in DEFAULT_FEATURE_ACCESS.items():
            if key in raw_access:
                store["feature_access"][key] = _normalize_feature_access_value(raw_access.get(key), default_value)

    raw_public = raw.get("public_features")
    if isinstance(raw_public, dict) and "liquidacion_final" not in (raw_access or {}):
        if "liquidacion_final_publica" in raw_public and bool(raw_public.get("liquidacion_final_publica")):
            store["feature_access"]["liquidacion_final"] = "public"

    store["feature_access"]["liquidacion_final"] = "public"
    store["feature_access"]["anexo_costo_empleador"] = "public"
    store["public_features"] = _feature_access_to_public_features(store["feature_access"])

    updated_at = raw.get("updated_at")
    if isinstance(updated_at, str) and updated_at.strip():
        store["updated_at"] = updated_at.strip()

    updated_by = raw.get("updated_by")
    if isinstance(updated_by, str):
        store["updated_by"] = updated_by.strip()

    return store


def _read_feature_store() -> Dict[str, Any]:
    if not ADMIN_FEATURES_FILE.exists():
        return _default_feature_store()

    try:
        raw = json.loads(ADMIN_FEATURES_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        raw = {}
    return _normalize_feature_store(raw)


def _write_feature_store(store: Dict[str, Any]) -> Dict[str, Any]:
    normalized = _normalize_feature_store(store)
    ADMIN_FEATURES_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ADMIN_FEATURES_FILE.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(ADMIN_FEATURES_FILE)
    return normalized


def _feature_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _clean_lead_text(value: Any, max_len: int = 300) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:max_len]


def _read_admin_leads() -> List[Dict[str, Any]]:
    if not ADMIN_LEADS_FILE.exists():
        return []

    try:
        raw = json.loads(ADMIN_LEADS_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        raw = []

    if not isinstance(raw, list):
        return []

    leads: List[Dict[str, Any]] = []
    for item in raw:
        if isinstance(item, dict):
            leads.append({
                "id": _clean_lead_text(item.get("id"), 80),
                "created_at": _clean_lead_text(item.get("created_at"), 80),
                "nombre": _clean_lead_text(item.get("nombre"), 160),
                "email": _clean_lead_text(item.get("email"), 220),
                "empresa": _clean_lead_text(item.get("empresa"), 180),
                "telefono": _clean_lead_text(item.get("telefono"), 80),
                "empleados": _clean_lead_text(item.get("empleados"), 80),
                "motivo": _clean_lead_text(item.get("motivo"), 220),
                "mensaje": _clean_lead_text(item.get("mensaje"), 500),
            })
    return leads


def _write_admin_leads(leads: List[Dict[str, Any]]) -> None:
    ADMIN_LEADS_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ADMIN_LEADS_FILE.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(leads[-2000:], ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(ADMIN_LEADS_FILE)


def _public_feature_payload(store: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "ok": True,
        "public_features": dict(store.get("public_features") or {}),
        "updated_at": store.get("updated_at"),
    }


def _admin_feature_payload(store: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "ok": True,
        "feature_access": dict(store.get("feature_access") or {}),
        "public_features": dict(store.get("public_features") or {}),
        "updated_at": store.get("updated_at"),
        "updated_by": store.get("updated_by") or "",
    }


def _require_admin_session(authorization: Optional[str]) -> Dict[str, Any]:
    return _read_admin_token(_extract_admin_token(authorization))


def _optional_admin_session(authorization: Optional[str]) -> Optional[Dict[str, Any]]:
    if not authorization:
        return None
    return _read_admin_token(_extract_admin_token(authorization))


def _is_public_feature_enabled(feature_name: str) -> bool:
    store = _read_feature_store()
    public_features = store.get("public_features") or {}
    return bool(public_features.get(feature_name))


def _get_feature_access(feature_name: str) -> str:
    store = _read_feature_store()
    feature_access = store.get("feature_access") or {}
    default_value = DEFAULT_FEATURE_ACCESS.get(feature_name, "off")
    return _normalize_feature_access_value(feature_access.get(feature_name), default_value)


def _require_admin_feature_access(authorization: Optional[str], feature_name: str) -> Dict[str, Any]:
    admin_payload = _require_admin_session(authorization)
    access = _get_feature_access(feature_name)
    if access not in {"admin_only", "public"}:
        raise HTTPException(status_code=403, detail="La función todavía no está habilitada en el panel.")
    return admin_payload


def _read_admin_companies() -> List[Dict[str, Any]]:
    if not ADMIN_COMPANIES_FILE.exists():
        return []

    try:
        raw = json.loads(ADMIN_COMPANIES_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        raw = []

    if not isinstance(raw, list):
        return []

    companies: List[Dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        razon_social = str(item.get("razon_social") or "").strip()
        if not razon_social:
            continue
        companies.append(
            {
                "id": str(item.get("id") or "").strip() or uuid.uuid4().hex[:12],
                "razon_social": razon_social,
                "cuit": str(item.get("cuit") or "").strip(),
                "rama": str(item.get("rama") or "").strip(),
                "email": str(item.get("email") or "").strip(),
                "telefono": str(item.get("telefono") or "").strip(),
                "actividad": str(item.get("actividad") or "").strip(),
                "logo_url": str(item.get("logo_url") or "").strip(),
                "direccion_calle": str(item.get("direccion_calle") or "").strip(),
                "direccion_numero": str(item.get("direccion_numero") or "").strip(),
                "direccion_piso": str(item.get("direccion_piso") or "").strip(),
                "direccion_depto": str(item.get("direccion_depto") or "").strip(),
                "localidad": str(item.get("localidad") or "").strip(),
                "provincia": str(item.get("provincia") or "").strip(),
                "codigo_postal": str(item.get("codigo_postal") or "").strip(),
                "estado": str(item.get("estado") or "prueba").strip() or "prueba",
                "observaciones": str(item.get("observaciones") or "").strip(),
                "created_at": str(item.get("created_at") or "").strip(),
                "updated_at": str(item.get("updated_at") or item.get("created_at") or "").strip(),
                "created_by": str(item.get("created_by") or "").strip(),
            }
        )
    return companies


def _write_admin_companies(companies: List[Dict[str, Any]]) -> None:
    ADMIN_COMPANIES_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ADMIN_COMPANIES_FILE.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(companies, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(ADMIN_COMPANIES_FILE)


def _read_admin_employees() -> List[Dict[str, Any]]:
    if not ADMIN_EMPLOYEES_FILE.exists():
        return []

    try:
        raw = json.loads(ADMIN_EMPLOYEES_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        raw = []

    if not isinstance(raw, list):
        return []

    employees: List[Dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        apellido_nombre = str(item.get("apellido_nombre") or "").strip()
        company_id = str(item.get("company_id") or "").strip()
        if not apellido_nombre or not company_id:
            continue
        employees.append(
            {
                "id": str(item.get("id") or "").strip() or uuid.uuid4().hex[:12],
                "company_id": company_id,
                "legajo": str(item.get("legajo") or "").strip(),
                "apellido_nombre": apellido_nombre,
                "cuil": str(item.get("cuil") or "").strip(),
                "sueldo_jornal": str(item.get("sueldo_jornal") or "").strip(),
                "categoria": str(item.get("categoria") or "").strip(),
                "tarea": str(item.get("tarea") or "").strip(),
                "fecha_ingreso": str(item.get("fecha_ingreso") or "").strip(),
                "obra_social": str(item.get("obra_social") or "").strip(),
                "obra_social_periodo": str(item.get("obra_social_periodo") or "").strip(),
                "deposito_previsional": str(item.get("deposito_previsional") or "").strip(),
                "deposito_previsional_fecha": str(item.get("deposito_previsional_fecha") or "").strip(),
                "lugar_pago": str(item.get("lugar_pago") or "").strip(),
                "estado": str(item.get("estado") or "prueba").strip() or "prueba",
                "observaciones": str(item.get("observaciones") or "").strip(),
                "created_at": str(item.get("created_at") or "").strip(),
                "updated_at": str(item.get("updated_at") or item.get("created_at") or "").strip(),
                "created_by": str(item.get("created_by") or "").strip(),
            }
        )
    return employees


def _write_admin_employees(employees: List[Dict[str, Any]]) -> None:
    ADMIN_EMPLOYEES_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ADMIN_EMPLOYEES_FILE.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(employees, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(ADMIN_EMPLOYEES_FILE)


def _read_json_list(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        return []
    return [item for item in raw if isinstance(item, dict)] if isinstance(raw, list) else []


def _write_json_list(path: Path, items: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def _digits(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _payroll_summary(result: Dict[str, Any]) -> Dict[str, float]:
    totals = result.get("totales") if isinstance(result, dict) else {}
    totals = totals if isinstance(totals, dict) else {}
    return {
        "rem": float(totals.get("rem") or 0),
        "nr": float(totals.get("nr") or 0),
        "ded": float(totals.get("ded") or 0),
        "neto": float(totals.get("neto") or 0),
    }


def _arca_validation(company_id: str, periodo: str) -> Dict[str, Any]:
    companies = _read_admin_companies()
    employees = _read_admin_employees()
    histories = _read_json_list(ADMIN_PAYROLL_HISTORY_FILE)
    mappings = _read_json_list(ADMIN_ARCA_MAPPINGS_FILE)
    company = next((item for item in companies if item.get("id") == company_id), None)
    selected = [
        item for item in histories
        if item.get("company_id") == company_id and item.get("periodo") == periodo and item.get("tipo") == "mensual"
    ]
    errors: List[str] = []
    warnings: List[str] = []
    if not company:
        errors.append("La empresa seleccionada no existe.")
    elif len(_digits(company.get("cuit"))) != 11:
        errors.append("La empresa necesita un CUIT valido de 11 digitos.")
    if not re.fullmatch(r"\d{6}", periodo or ""):
        errors.append("El periodo debe tener formato AAAAMM.")
    if not selected:
        errors.append("No hay liquidaciones mensuales guardadas para el periodo.")

    employee_by_id = {str(item.get("id") or ""): item for item in employees}
    mapped_names = {
        str(item.get("concepto") or "").strip().lower()
        for item in mappings if item.get("company_id") == company_id
    }
    missing_concepts = set()
    for history in selected:
        employee = employee_by_id.get(str(history.get("employee_id") or ""))
        if not employee:
            errors.append(f"La liquidacion {history.get('id')} no tiene un empleado vigente.")
            continue
        if len(_digits(employee.get("cuil"))) != 11:
            errors.append(f"{employee.get('apellido_nombre')}: falta un CUIL valido de 11 digitos.")
        result = history.get("resultado") if isinstance(history.get("resultado"), dict) else {}
        for concept in result.get("items") or []:
            if not isinstance(concept, dict):
                continue
            name = str(concept.get("concepto") or "").strip()
            amount = float(concept.get("r") or 0) + float(concept.get("n") or 0) + float(concept.get("d") or 0)
            if name and amount and name.lower() not in mapped_names:
                missing_concepts.add(name)
    if missing_concepts:
        errors.append(f"Faltan parametrizar {len(missing_concepts)} conceptos para ARCA.")
    warnings.append("El TXT se habilitara cuando tambien esten completos los datos F.931 de cada trabajador.")
    return {
        "ok": not errors,
        "ready_for_txt": False,
        "company_id": company_id,
        "periodo": periodo,
        "liquidaciones": len(selected),
        "errors": errors,
        "warnings": warnings,
        "missing_concepts": sorted(missing_concepts),
    }


def _default_admin_company_state() -> Dict[str, Any]:
    return {
        "active_company_id": "",
        "updated_at": None,
        "updated_by": "",
    }


def _normalize_admin_company_state(raw: Any) -> Dict[str, Any]:
    state = _default_admin_company_state()
    if not isinstance(raw, dict):
        return state

    active_company_id = raw.get("active_company_id")
    if isinstance(active_company_id, str):
        state["active_company_id"] = active_company_id.strip()

    updated_at = raw.get("updated_at")
    if isinstance(updated_at, str) and updated_at.strip():
        state["updated_at"] = updated_at.strip()

    updated_by = raw.get("updated_by")
    if isinstance(updated_by, str):
        state["updated_by"] = updated_by.strip()

    return state


def _read_admin_company_state() -> Dict[str, Any]:
    if not ADMIN_COMPANY_STATE_FILE.exists():
        return _default_admin_company_state()

    try:
        raw = json.loads(ADMIN_COMPANY_STATE_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        raw = {}
    return _normalize_admin_company_state(raw)


def _write_admin_company_state(state: Dict[str, Any]) -> Dict[str, Any]:
    normalized = _normalize_admin_company_state(state)
    ADMIN_COMPANY_STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ADMIN_COMPANY_STATE_FILE.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(ADMIN_COMPANY_STATE_FILE)
    return normalized


def _resolve_active_admin_company_id(companies: List[Dict[str, Any]], requested_id: str) -> str:
    wanted = str(requested_id or "").strip()
    if wanted and any(str(item.get("id") or "") == wanted for item in companies):
        return wanted
    if companies:
        return str(companies[0].get("id") or "").strip()
    return ""


def _default_admin_employee_state() -> Dict[str, Any]:
    return {
        "active_employee_by_company_id": {},
        "updated_at": None,
        "updated_by": "",
    }


def _normalize_admin_employee_state(raw: Any) -> Dict[str, Any]:
    state = _default_admin_employee_state()
    if not isinstance(raw, dict):
        return state

    raw_active = raw.get("active_employee_by_company_id")
    if isinstance(raw_active, dict):
        state["active_employee_by_company_id"] = {
            str(company_id).strip(): str(employee_id).strip()
            for company_id, employee_id in raw_active.items()
            if str(company_id).strip() and str(employee_id).strip()
        }

    updated_at = raw.get("updated_at")
    if isinstance(updated_at, str) and updated_at.strip():
        state["updated_at"] = updated_at.strip()

    updated_by = raw.get("updated_by")
    if isinstance(updated_by, str):
        state["updated_by"] = updated_by.strip()

    return state


def _read_admin_employee_state() -> Dict[str, Any]:
    if not ADMIN_EMPLOYEE_STATE_FILE.exists():
        return _default_admin_employee_state()

    try:
        raw = json.loads(ADMIN_EMPLOYEE_STATE_FILE.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        raw = {}
    return _normalize_admin_employee_state(raw)


def _write_admin_employee_state(state: Dict[str, Any]) -> Dict[str, Any]:
    normalized = _normalize_admin_employee_state(state)
    ADMIN_EMPLOYEE_STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ADMIN_EMPLOYEE_STATE_FILE.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(ADMIN_EMPLOYEE_STATE_FILE)
    return normalized


def _resolve_active_admin_employee_id(
    employees: List[Dict[str, Any]],
    company_id: str,
    requested_id: str,
) -> str:
    company = str(company_id or "").strip()
    wanted = str(requested_id or "").strip()
    company_employees = [item for item in employees if str(item.get("company_id") or "") == company]
    if wanted and any(str(item.get("id") or "") == wanted for item in company_employees):
        return wanted
    if company_employees:
        return str(company_employees[0].get("id") or "").strip()
    return ""


def _employee_import_key(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _employee_import_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _employee_import_digits(value: Any) -> str:
    return re.sub(r"\D+", "", _employee_import_text(value))


def _employee_import_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return str(value.isoformat())[:10]
    text = _employee_import_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _catalog_value(value: Any, candidates: List[str]) -> str:
    wanted = _employee_import_key(value)
    for candidate in candidates:
        if _employee_import_key(candidate) == wanted:
            return candidate
    return ""


def _parse_employee_import(contents: bytes) -> Dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel .xlsx válido.") from error

    sheet = workbook.worksheets[0]
    iterator = sheet.iter_rows(values_only=True)
    raw_headers = next(iterator, None)
    if not raw_headers:
        raise HTTPException(status_code=400, detail="La planilla no contiene encabezados.")

    columns: Dict[int, str] = {}
    for index, header in enumerate(raw_headers):
        mapped = EMPLOYEE_IMPORT_HEADERS.get(_employee_import_key(header))
        if mapped:
            columns[index] = mapped
    required = {"full_name", "cuil", "rama", "agrup", "category", "start_date"}
    missing = sorted(required.difference(columns.values()))
    if missing:
        raise HTTPException(
            status_code=400,
            detail="Faltan columnas obligatorias: apellido_y_nombre, cuil, fecha_ingreso, rama, agrupamiento y categoria.",
        )

    meta = get_meta()
    ramas = list(meta.get("ramas") or [])
    agrupamientos = dict(meta.get("agrupamientos") or {})
    categorias = dict(meta.get("categorias") or {})
    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    seen_cuils: set[str] = set()

    for excel_row, raw_row in enumerate(iterator, start=2):
        if excel_row > EMPLOYEE_IMPORT_MAX_ROWS + 1:
            raise HTTPException(status_code=400, detail=f"La importación admite hasta {EMPLOYEE_IMPORT_MAX_ROWS} empleados por archivo.")
        values = {field: raw_row[index] if index < len(raw_row) else None for index, field in columns.items()}
        if not any(_employee_import_text(value) for value in values.values()):
            continue

        row_errors: List[str] = []
        full_name = _employee_import_text(values.get("full_name"))
        cuil = _employee_import_digits(values.get("cuil"))
        rama = _catalog_value(values.get("rama"), ramas)
        agrup = _catalog_value(values.get("agrup"), list(agrupamientos.get(rama) or [])) if rama else ""
        category = _catalog_value(values.get("category"), list((categorias.get(rama) or {}).get(agrup) or [])) if agrup else ""
        start_date = _employee_import_date(values.get("start_date"))
        status_key = _employee_import_key(values.get("status") or "active")
        status = {"activo": "active", "active": "active", "pausado": "paused", "paused": "paused", "desvinculado": "terminated", "terminated": "terminated"}.get(status_key, "")

        if not full_name:
            row_errors.append("Falta apellido y nombre.")
        if len(cuil) != 11:
            row_errors.append("El CUIL debe tener 11 dígitos.")
        elif cuil in seen_cuils:
            row_errors.append("El CUIL está repetido dentro de la planilla.")
        if not rama:
            row_errors.append("La rama no existe en el catálogo salarial.")
        if rama and not agrup:
            row_errors.append("El agrupamiento no corresponde a la rama.")
        if agrup and not category:
            row_errors.append("La categoría no corresponde al agrupamiento.")
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", start_date):
            row_errors.append("La fecha de ingreso no es válida.")
        if not status:
            row_errors.append("El estado debe ser Activo, Pausado o Desvinculado.")
        cbu = _employee_import_digits(values.get("cbu"))
        if cbu and len(cbu) != 22:
            row_errors.append("El CBU debe tener 22 dígitos.")

        if cuil:
            seen_cuils.add(cuil)
        if row_errors:
            errors.append({"row": excel_row, "employee": full_name, "messages": row_errors})
            continue

        weekly_hours_text = _employee_import_text(values.get("weekly_hours"))
        weekly_hours = float(weekly_hours_text.replace(",", ".")) if re.fullmatch(r"\d+(?:[\.,]\d+)?", weekly_hours_text) else ""
        if weekly_hours == "":
            warnings.append({"row": excel_row, "message": "No se informaron horas semanales; se utilizarán 48 horas."})
            weekly_hours = 48

        rows.append({
            "file_number": _employee_import_text(values.get("file_number")),
            "full_name": full_name,
            "cuil": cuil,
            "category": category,
            "start_date": start_date,
            "status": status,
            "profile": {
                "dni": _employee_import_digits(values.get("dni")),
                "birth_date": _employee_import_date(values.get("birth_date")),
                "end_date": _employee_import_date(values.get("end_date")),
                "rama": rama,
                "agrup": agrup,
                "agreement": {
                    "GENERAL": "CCT 130/75", "CEREALES": "CCT 130/75", "FÚNEBRES": "CCT 177/75",
                    "AGUA POTABLE": "CCT Agua potable", "CALL CENTER": "CCT 781/20", "TURISMO": "CCT 547/08",
                }.get(rama, ""),
                "workday": _employee_import_text(values.get("workday")),
                "weekly_hours": weekly_hours,
                "contract_modality": _employee_import_text(values.get("contract_modality")),
                "email": _employee_import_text(values.get("email")).lower(),
                "phone": _employee_import_text(values.get("phone")),
                "address": _employee_import_text(values.get("address")),
                "locality": _employee_import_text(values.get("locality")),
                "province": _employee_import_text(values.get("province")),
                "health_insurance_name": _employee_import_text(values.get("health_insurance_name")),
                "health_insurance_member": _employee_import_text(values.get("health_insurance_member")),
                "health_insurance_start": _employee_import_date(values.get("health_insurance_start")),
                "union_name": _employee_import_text(values.get("union_name")),
                "union_member": _employee_import_text(values.get("union_member")),
                "bank_name": _employee_import_text(values.get("bank_name")),
                "cbu": cbu,
                "bank_alias": _employee_import_text(values.get("bank_alias")),
                "payment_method": _employee_import_text(values.get("payment_method")) or "Transferencia bancaria",
            },
            "arca_profile": {
                "health_insurance_code": _employee_import_digits(values.get("health_insurance_code")),
                "contract_modality_code": _employee_import_digits(values.get("contract_modality_code")),
                "union_code": _employee_import_digits(values.get("union_code")),
            },
        })

    workbook.close()
    return {"ok": not errors, "rows": rows, "errors": errors, "warnings": warnings, "count": len(rows)}


# ========= HOME → HTML =========
@app.get("/", include_in_schema=False)
def home():
    if PUBLIC_STATIC_INDEX_FILE.exists():
        return FileResponse(PUBLIC_STATIC_INDEX_FILE, headers=NOINDEX_HEADERS)

    if PUBLIC_INDEX_FILE.exists():
        return FileResponse(PUBLIC_INDEX_FILE, headers=NOINDEX_HEADERS)

    if ADMIN_INDEX_FILE.exists():
        return FileResponse(ADMIN_INDEX_FILE, headers=NOINDEX_HEADERS)

    return {"ok": True, "error": "index.html no encontrado"}


@app.get("/admin/app", include_in_schema=False)
def admin_app(admin_token: str = Query(default="")):
    if admin_token:
        _read_admin_token(admin_token)
    if PUBLIC_STATIC_INDEX_FILE.exists():
        return FileResponse(PUBLIC_STATIC_INDEX_FILE, headers=NOINDEX_HEADERS)
    if ADMIN_INDEX_FILE.exists():
        return FileResponse(ADMIN_INDEX_FILE, headers=NOINDEX_HEADERS)
    return HTMLResponse("<h1>Panel administrador no encontrado</h1>", status_code=404, headers=NOINDEX_HEADERS)


@app.get("/empresas", include_in_schema=False)
@app.get("/empresas/", include_in_schema=False)
def company_portal():
    if COMPANY_PORTAL_FILE.exists():
        return FileResponse(COMPANY_PORTAL_FILE, headers=NOINDEX_HEADERS)
    return HTMLResponse("<h1>Portal de empresas no encontrado</h1>", status_code=404, headers=NOINDEX_HEADERS)


@app.get("/plantilla-importacion-empleados.xlsx", include_in_schema=False)
def employee_import_template():
    if EMPLOYEE_IMPORT_TEMPLATE_FILE.exists():
        return FileResponse(
            EMPLOYEE_IMPORT_TEMPLATE_FILE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="plantilla_importacion_empleados.xlsx",
            headers=NOINDEX_HEADERS,
        )
    raise HTTPException(status_code=404, detail="Plantilla no disponible.")


@app.get("/manual-liquidacion-recibos-libro-sueldos-digital-arca.pdf", include_in_schema=False)
def company_payroll_manual():
    if COMPANY_PAYROLL_MANUAL_FILE.exists():
        return FileResponse(
            COMPANY_PAYROLL_MANUAL_FILE,
            media_type="application/pdf",
            headers=NOINDEX_HEADERS,
        )
    raise HTTPException(status_code=404, detail="Manual no disponible.")


@app.post("/employees-import-preview")
async def employee_import_preview(request: Request, file: UploadFile = File(...)):
    _enforce_rate_limit(request, "employee-import", 10, 600)
    filename = str(file.filename or "")
    if Path(filename).suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Seleccioná un archivo Excel con extensión .xlsx.")
    contents = await file.read(EMPLOYEE_IMPORT_MAX_BYTES + 1)
    if len(contents) > EMPLOYEE_IMPORT_MAX_BYTES:
        raise HTTPException(status_code=413, detail="El archivo supera el máximo permitido de 3 MB.")
    return _parse_employee_import(contents)

# ========= HEALTH =========
@app.get("/health")
def health():
    return {"ok": True, "servicio": "motor-sueldos-faecys"}


@app.post("/leads")
def create_public_lead(payload: PublicLeadCreate, request: Request):
    _enforce_rate_limit(request, "public-leads", 5, 3600)
    nombre = _clean_lead_text(payload.nombre, 160)
    email = _clean_lead_text(payload.email, 220)
    empresa = _clean_lead_text(payload.empresa, 180)
    telefono = _clean_lead_text(payload.telefono, 80)
    empleados = _clean_lead_text(payload.empleados, 80)
    motivo = _clean_lead_text(payload.motivo, 220) or "Consulta empresa / nomina"
    mensaje = _clean_lead_text(payload.mensaje, 500)

    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre es obligatorio.")
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="El email es obligatorio.")
    if not empresa:
        raise HTTPException(status_code=400, detail="La empresa es obligatoria.")

    lead = {
        "id": uuid.uuid4().hex[:12],
        "created_at": _feature_timestamp(),
        "nombre": nombre,
        "email": email,
        "empresa": empresa,
        "telefono": telefono,
        "empleados": empleados,
        "motivo": motivo,
        "mensaje": mensaje,
    }
    leads = _read_admin_leads()
    leads.append(lead)
    _write_admin_leads(leads)
    return {"ok": True, "id": lead["id"]}


@app.post("/admin/login")
def admin_login(payload: AdminLoginRequest, request: Request):
    _require_admin_security_config()
    _enforce_rate_limit(request, "admin-login", 8, 900)
    email = payload.email.strip().lower()
    password = payload.password

    valid_email = hmac.compare_digest(email, ADMIN_LOGIN_EMAIL)
    valid_password = hmac.compare_digest(password, ADMIN_LOGIN_PASSWORD)

    if not (valid_email and valid_password):
        raise HTTPException(status_code=401, detail="Credenciales de administrador inválidas.")

    return {
        "ok": True,
        "token": _issue_admin_token(email),
        "email": ADMIN_LOGIN_EMAIL,
        "role": "admin",
        "expires_in": ADMIN_TOKEN_TTL_SECONDS,
    }


@app.get("/admin/session")
def admin_session(authorization: Optional[str] = Header(default=None)):
    payload = _require_admin_session(authorization)
    return {
        "ok": True,
        "authenticated": True,
        "role": payload["role"],
        "email": payload["email"],
        "expires_at": payload["exp"],
    }


@app.get("/features")
def public_features():
    store = _read_feature_store()
    return _public_feature_payload(store)


@app.get("/admin/features")
def admin_features(authorization: Optional[str] = Header(default=None)):
    _require_admin_session(authorization)
    store = _read_feature_store()
    return _admin_feature_payload(store)


@app.get("/admin/leads")
def admin_leads(authorization: Optional[str] = Header(default=None)):
    _require_admin_session(authorization)
    leads = sorted(_read_admin_leads(), key=lambda item: item.get("created_at") or "", reverse=True)
    return {
        "ok": True,
        "count": len(leads),
        "items": leads,
    }


@app.put("/admin/features")
def update_admin_features(payload: AdminFeaturesUpdate, authorization: Optional[str] = Header(default=None)):
    admin_payload = _require_admin_session(authorization)
    store = _read_feature_store()
    feature_access = dict(store.get("feature_access") or {})
    public_features = dict(store.get("public_features") or {})

    for feature_name in DEFAULT_FEATURE_ACCESS:
        next_value = getattr(payload, feature_name, None)
        if next_value is not None:
            feature_access[feature_name] = _normalize_feature_access_value(next_value, DEFAULT_FEATURE_ACCESS[feature_name])

    if payload.liquidacion_final_publica is not None and payload.liquidacion_final is None:
        feature_access["liquidacion_final"] = "public" if bool(payload.liquidacion_final_publica) else "admin_only"

    feature_access["liquidacion_final"] = "public"
    feature_access["anexo_costo_empleador"] = "public"
    public_features = _feature_access_to_public_features(feature_access)
    store["feature_access"] = feature_access
    store["public_features"] = public_features
    store["updated_at"] = _feature_timestamp()
    store["updated_by"] = str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower()

    try:
        saved = _write_feature_store(store)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar la configuraciÃ³n del panel.") from exc
    return _admin_feature_payload(saved)


@app.get("/admin/companies")
def admin_companies(authorization: Optional[str] = Header(default=None)):
    _require_admin_feature_access(authorization, "registro_empresas")
    items = sorted(_read_admin_companies(), key=lambda item: (item.get("razon_social") or "").lower())
    state = _read_admin_company_state()
    active_company_id = _resolve_active_admin_company_id(items, state.get("active_company_id") or "")
    return {
        "ok": True,
        "items": items,
        "count": len(items),
        "active_company_id": active_company_id,
        "active_company_updated_at": state.get("updated_at"),
        "active_company_updated_by": state.get("updated_by") or "",
    }


@app.post("/admin/companies")
def create_admin_company(payload: AdminCompanyCreate, authorization: Optional[str] = Header(default=None)):
    admin_payload = _require_admin_feature_access(authorization, "registro_empresas")
    admin_email = str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower()
    razon_social = payload.razon_social.strip()
    if not razon_social:
        raise HTTPException(status_code=400, detail="La razón social es obligatoria.")

    companies = _read_admin_companies()
    now = _feature_timestamp()
    company = {
        "id": uuid.uuid4().hex[:12],
        "razon_social": razon_social,
        "cuit": payload.cuit.strip(),
        "rama": payload.rama.strip(),
        "email": payload.email.strip(),
        "telefono": payload.telefono.strip(),
        "actividad": payload.actividad.strip(),
        "logo_url": payload.logo_url.strip(),
        "direccion_calle": payload.direccion_calle.strip(),
        "direccion_numero": payload.direccion_numero.strip(),
        "direccion_piso": payload.direccion_piso.strip(),
        "direccion_depto": payload.direccion_depto.strip(),
        "localidad": payload.localidad.strip(),
        "provincia": payload.provincia.strip(),
        "codigo_postal": payload.codigo_postal.strip(),
        "estado": payload.estado.strip() or "prueba",
        "observaciones": payload.observaciones.strip(),
        "created_at": now,
        "updated_at": now,
        "created_by": admin_email,
    }
    companies.append(company)
    company_state = {
        "active_company_id": company["id"],
        "updated_at": now,
        "updated_by": admin_email,
    }

    try:
        _write_admin_companies(companies)
        _write_admin_company_state(company_state)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar la empresa de prueba.") from exc

    return {
        "ok": True,
        "item": company,
        "count": len(companies),
        "active_company_id": company["id"],
    }


@app.put("/admin/companies/active")
def set_active_admin_company(payload: AdminCompanyActiveUpdate, authorization: Optional[str] = Header(default=None)):
    admin_payload = _require_admin_feature_access(authorization, "registro_empresas")
    companies = sorted(_read_admin_companies(), key=lambda item: (item.get("razon_social") or "").lower())
    requested_id = payload.company_id.strip()

    if requested_id and not any(str(item.get("id") or "") == requested_id for item in companies):
        raise HTTPException(status_code=404, detail="La empresa seleccionada no existe.")

    state = {
        "active_company_id": requested_id,
        "updated_at": _feature_timestamp(),
        "updated_by": str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower(),
    }

    try:
        saved = _write_admin_company_state(state)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar la empresa activa.") from exc

    return {
        "ok": True,
        "active_company_id": saved.get("active_company_id") or "",
        "updated_at": saved.get("updated_at"),
        "updated_by": saved.get("updated_by") or "",
        "count": len(companies),
    }


@app.get("/admin/employees")
def admin_employees(company_id: str = Query(default=""), authorization: Optional[str] = Header(default=None)):
    _require_admin_feature_access(authorization, "registro_empresas")
    companies = sorted(_read_admin_companies(), key=lambda item: (item.get("razon_social") or "").lower())
    company_state = _read_admin_company_state()
    selected_company_id = str(company_id or "").strip() or _resolve_active_admin_company_id(
        companies,
        company_state.get("active_company_id") or "",
    )

    if selected_company_id and not any(str(item.get("id") or "") == selected_company_id for item in companies):
        raise HTTPException(status_code=404, detail="La empresa seleccionada no existe.")

    employees = sorted(
        [
            item
            for item in _read_admin_employees()
            if not selected_company_id or str(item.get("company_id") or "") == selected_company_id
        ],
        key=lambda item: ((item.get("apellido_nombre") or "").lower(), item.get("legajo") or ""),
    )
    employee_state = _read_admin_employee_state()
    active_map = employee_state.get("active_employee_by_company_id") or {}
    active_employee_id = _resolve_active_admin_employee_id(
        employees,
        selected_company_id,
        active_map.get(selected_company_id) or "",
    )
    return {
        "ok": True,
        "company_id": selected_company_id,
        "items": employees,
        "count": len(employees),
        "active_employee_id": active_employee_id,
        "active_employee_updated_at": employee_state.get("updated_at"),
        "active_employee_updated_by": employee_state.get("updated_by") or "",
    }


@app.post("/admin/employees")
def create_admin_employee(payload: AdminEmployeeCreate, authorization: Optional[str] = Header(default=None)):
    admin_payload = _require_admin_feature_access(authorization, "registro_empresas")
    admin_email = str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower()
    company_id = payload.company_id.strip()
    apellido_nombre = payload.apellido_nombre.strip()

    if not company_id:
        raise HTTPException(status_code=400, detail="Seleccioná una empresa para cargar el empleado.")
    if not any(str(item.get("id") or "") == company_id for item in _read_admin_companies()):
        raise HTTPException(status_code=404, detail="La empresa seleccionada no existe.")
    if not apellido_nombre:
        raise HTTPException(status_code=400, detail="El apellido y nombre del empleado es obligatorio.")

    employees = _read_admin_employees()
    now = _feature_timestamp()
    employee = {
        "id": uuid.uuid4().hex[:12],
        "company_id": company_id,
        "legajo": payload.legajo.strip(),
        "apellido_nombre": apellido_nombre,
        "cuil": payload.cuil.strip(),
        "sueldo_jornal": payload.sueldo_jornal.strip(),
        "categoria": payload.categoria.strip(),
        "tarea": payload.tarea.strip(),
        "fecha_ingreso": payload.fecha_ingreso.strip(),
        "obra_social": payload.obra_social.strip(),
        "obra_social_periodo": payload.obra_social_periodo.strip(),
        "deposito_previsional": payload.deposito_previsional.strip(),
        "deposito_previsional_fecha": payload.deposito_previsional_fecha.strip(),
        "lugar_pago": payload.lugar_pago.strip(),
        "estado": payload.estado.strip() or "prueba",
        "observaciones": payload.observaciones.strip(),
        "created_at": now,
        "updated_at": now,
        "created_by": admin_email,
    }
    employees.append(employee)

    employee_state = _read_admin_employee_state()
    active_map = dict(employee_state.get("active_employee_by_company_id") or {})
    active_map[company_id] = employee["id"]
    employee_state["active_employee_by_company_id"] = active_map
    employee_state["updated_at"] = now
    employee_state["updated_by"] = admin_email

    try:
        _write_admin_employees(employees)
        _write_admin_employee_state(employee_state)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar el empleado de prueba.") from exc

    return {
        "ok": True,
        "item": employee,
        "count": len([item for item in employees if item.get("company_id") == company_id]),
        "active_employee_id": employee["id"],
    }


@app.put("/admin/employees/active")
def set_active_admin_employee(payload: AdminEmployeeActiveUpdate, authorization: Optional[str] = Header(default=None)):
    admin_payload = _require_admin_feature_access(authorization, "registro_empresas")
    company_id = payload.company_id.strip()
    employee_id = payload.employee_id.strip()

    if not company_id:
        raise HTTPException(status_code=400, detail="Seleccioná una empresa para elegir empleado.")
    if not any(str(item.get("id") or "") == company_id for item in _read_admin_companies()):
        raise HTTPException(status_code=404, detail="La empresa seleccionada no existe.")

    employees = [
        item
        for item in _read_admin_employees()
        if str(item.get("company_id") or "") == company_id
    ]
    if employee_id and not any(str(item.get("id") or "") == employee_id for item in employees):
        raise HTTPException(status_code=404, detail="El empleado seleccionado no existe.")

    employee_state = _read_admin_employee_state()
    active_map = dict(employee_state.get("active_employee_by_company_id") or {})
    if employee_id:
        active_map[company_id] = employee_id
    else:
        active_map.pop(company_id, None)
    employee_state["active_employee_by_company_id"] = active_map
    employee_state["updated_at"] = _feature_timestamp()
    employee_state["updated_by"] = str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower()

    try:
        saved = _write_admin_employee_state(employee_state)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar el empleado activo.") from exc

    saved_map = saved.get("active_employee_by_company_id") or {}
    return {
        "ok": True,
        "company_id": company_id,
        "active_employee_id": saved_map.get(company_id) or "",
        "updated_at": saved.get("updated_at"),
        "updated_by": saved.get("updated_by") or "",
        "count": len(employees),
    }


@app.get("/admin/payroll-history")
def admin_payroll_history(
    company_id: str = Query(default=""),
    periodo: str = Query(default=""),
    authorization: Optional[str] = Header(default=None),
):
    _require_admin_feature_access(authorization, "gestion_nomina")
    company = company_id.strip()
    period = re.sub(r"\D+", "", periodo or "")[:6]
    items = [
        item for item in _read_json_list(ADMIN_PAYROLL_HISTORY_FILE)
        if (not company or item.get("company_id") == company)
        and (not period or item.get("periodo") == period)
    ]
    items.sort(key=lambda item: (item.get("periodo") or "", item.get("updated_at") or ""), reverse=True)
    return {"ok": True, "count": len(items), "items": items}


@app.post("/admin/payroll-history")
def save_admin_payroll_history(
    payload: AdminPayrollHistoryCreate,
    authorization: Optional[str] = Header(default=None),
):
    admin_payload = _require_admin_feature_access(authorization, "gestion_nomina")
    company_id = payload.company_id.strip()
    employee_id = payload.employee_id.strip()
    periodo = re.sub(r"\D+", "", payload.periodo or "")[:6]
    tipo = payload.tipo.strip().lower() or "mensual"
    if not re.fullmatch(r"\d{6}", periodo):
        raise HTTPException(status_code=400, detail="El periodo debe tener formato AAAAMM.")
    if not any(item.get("id") == company_id for item in _read_admin_companies()):
        raise HTTPException(status_code=404, detail="La empresa seleccionada no existe.")
    employee = next(
        (item for item in _read_admin_employees() if item.get("id") == employee_id and item.get("company_id") == company_id),
        None,
    )
    if not employee:
        raise HTTPException(status_code=404, detail="El empleado seleccionado no pertenece a la empresa.")
    if not isinstance(payload.resultado, dict) or not isinstance(payload.resultado.get("items"), list):
        raise HTTPException(status_code=400, detail="Primero realiza una liquidacion valida.")

    histories = _read_json_list(ADMIN_PAYROLL_HISTORY_FILE)
    now = _feature_timestamp()
    existing = next(
        (
            item for item in histories
            if item.get("company_id") == company_id
            and item.get("employee_id") == employee_id
            and item.get("periodo") == periodo
            and item.get("tipo") == tipo
        ),
        None,
    )
    record_id = str(existing.get("id") or "") if existing else uuid.uuid4().hex[:16]
    record = {
        "id": record_id,
        "company_id": company_id,
        "employee_id": employee_id,
        "employee_name": employee.get("apellido_nombre") or "",
        "employee_cuil": employee.get("cuil") or "",
        "periodo": periodo,
        "tipo": tipo,
        "inputs": payload.inputs if isinstance(payload.inputs, dict) else {},
        "resultado": payload.resultado,
        "resumen": _payroll_summary(payload.resultado),
        "created_at": existing.get("created_at") if existing else now,
        "updated_at": now,
        "updated_by": str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower(),
    }
    if existing:
        histories = [record if item is existing else item for item in histories]
    else:
        histories.append(record)
    try:
        _write_json_list(ADMIN_PAYROLL_HISTORY_FILE, histories)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar la liquidacion en el historial.") from exc
    return {"ok": True, "item": record, "updated": bool(existing)}


@app.delete("/admin/payroll-history/{record_id}")
def delete_admin_payroll_history(record_id: str, authorization: Optional[str] = Header(default=None)):
    _require_admin_feature_access(authorization, "gestion_nomina")
    histories = _read_json_list(ADMIN_PAYROLL_HISTORY_FILE)
    next_items = [item for item in histories if str(item.get("id") or "") != record_id]
    if len(next_items) == len(histories):
        raise HTTPException(status_code=404, detail="La liquidacion guardada no existe.")
    _write_json_list(ADMIN_PAYROLL_HISTORY_FILE, next_items)
    return {"ok": True, "count": len(next_items)}


@app.get("/admin/arca/mappings")
def admin_arca_mappings(company_id: str = Query(default=""), authorization: Optional[str] = Header(default=None)):
    _require_admin_feature_access(authorization, "gestion_nomina")
    company = company_id.strip()
    items = [item for item in _read_json_list(ADMIN_ARCA_MAPPINGS_FILE) if not company or item.get("company_id") == company]
    items.sort(key=lambda item: str(item.get("concepto") or "").lower())
    return {"ok": True, "count": len(items), "items": items}


@app.post("/admin/arca/mappings")
def save_admin_arca_mapping(payload: AdminArcaMappingCreate, authorization: Optional[str] = Header(default=None)):
    admin_payload = _require_admin_feature_access(authorization, "gestion_nomina")
    company_id = payload.company_id.strip()
    concepto = payload.concepto.strip()
    employer_code = re.sub(r"[^A-Za-z0-9_-]+", "", payload.codigo_empleador.strip())[:10]
    arca_code = _digits(payload.codigo_arca)[:6]
    if not any(item.get("id") == company_id for item in _read_admin_companies()):
        raise HTTPException(status_code=404, detail="La empresa seleccionada no existe.")
    if not concepto or not employer_code or len(arca_code) != 6:
        raise HTTPException(status_code=400, detail="Completa concepto, codigo empleador y codigo ARCA de 6 digitos.")
    mappings = _read_json_list(ADMIN_ARCA_MAPPINGS_FILE)
    now = _feature_timestamp()
    record = {
        "id": uuid.uuid4().hex[:16],
        "company_id": company_id,
        "concepto": concepto,
        "codigo_empleador": employer_code,
        "codigo_arca": arca_code,
        "unidad": (payload.unidad.strip() or "$")[:1],
        "updated_at": now,
        "updated_by": str(admin_payload.get("email") or ADMIN_LOGIN_EMAIL).strip().lower(),
    }
    replaced = False
    for index, item in enumerate(mappings):
        if item.get("company_id") == company_id and str(item.get("concepto") or "").strip().lower() == concepto.lower():
            record["id"] = item.get("id") or record["id"]
            mappings[index] = record
            replaced = True
            break
    if not replaced:
        mappings.append(record)
    _write_json_list(ADMIN_ARCA_MAPPINGS_FILE, mappings)
    return {"ok": True, "item": record, "updated": replaced}


@app.get("/admin/arca/validate")
def validate_admin_arca(
    company_id: str = Query(default=""),
    periodo: str = Query(default=""),
    authorization: Optional[str] = Header(default=None),
):
    _require_admin_feature_access(authorization, "gestion_nomina")
    return _arca_validation(company_id.strip(), re.sub(r"\D+", "", periodo or "")[:6])


@app.post("/admin/companies/logo")
async def upload_admin_company_logo(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
):
    _require_admin_feature_access(authorization, "registro_empresas")
    original_name = str(file.filename or "").strip()
    suffix = Path(original_name).suffix.lower()

    if suffix not in ADMIN_COMPANY_ASSET_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Formato de logo no valido. Usa PNG, JPG, WEBP, GIF o SVG.",
        )

    raw = await file.read()
    await file.close()

    if not raw:
        raise HTTPException(status_code=400, detail="El archivo de logo esta vacio.")
    if len(raw) > ADMIN_COMPANY_ASSET_MAX_BYTES:
        raise HTTPException(status_code=400, detail="El logo supera el tamano maximo permitido.")

    ADMIN_COMPANY_ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    stem = _sanitize_admin_asset_stem(Path(original_name).stem)
    asset_name = f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{stem}_{uuid.uuid4().hex[:8]}{suffix}"
    asset_path = ADMIN_COMPANY_ASSETS_DIR / asset_name

    try:
        asset_path.write_bytes(raw)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="No se pudo guardar el logo de la empresa.") from exc

    return {
        "ok": True,
        "file_name": asset_name,
        "original_name": original_name,
        "logo_url": _build_admin_company_asset_url(asset_name),
    }


@app.get("/admin/company-assets/{asset_name}")
def admin_company_asset(asset_name: str):
    safe_name = Path(asset_name).name
    if safe_name != asset_name:
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")

    asset_path = ADMIN_COMPANY_ASSETS_DIR / safe_name
    if not asset_path.is_file():
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")

    media_type = mimetypes.guess_type(str(asset_path))[0] or "application/octet-stream"
    return FileResponse(asset_path, media_type=media_type)

# ========= META =========
@app.get("/meta")
def meta():
    return get_meta()

# ========= PAYLOAD (bases del maestro) =========
@app.get("/payload")
def payload(
    rama: str,
    mes: str,
    agrup: str = "—",
    categoria: str = "—",
    conex_cat: str = "",
    conexiones: int = 0,
):
    return get_payload(
        rama=rama,
        mes=mes,
        agrup=agrup,
        categoria=categoria,
        conex_cat=conex_cat,
        conexiones=conexiones,
    )

# ========= CALCULAR (recibo completo) =========
@app.get("/calcular")
def calcular(
    rama: str,
    agrup: str,
    categoria: str,
    mes: str,
    jornada: float = 48.0,
    basico_manual: float = 0,
    anios_antig: float = 0,
    osecac: bool = True,
    obra_social_sobre_no_rem: bool = True,
    afiliado: bool = False,
    sind_pct: float = 0,
    sind_fijo: float = 0,
    titulo_pct: float = 0,
    zona_pct: float = 0,
    fer_no_trab: int = 0,
    fer_trab: int = 0,
    vac_goz: int = 0,
    aus_inj: int = 0,
    jubilado: bool = False,
    susp_dias: int = 0,
    embargo: float = 0,
    # Horas
    hex50: float = 0,
    hex100: float = 0,
    hs_noct: float = 0,
    # KM (Chofer/Ayudante)
    km_tipo: str = "",
    km_menos100: float = 0,
    km_mas100: float = 0,
    # Etapa 5/6: A cuenta (REM) / Viáticos (NR sin aportes)
    a_cuenta_rem: float = 0,
    viaticos_nr: float = 0,

    # Etapa 7: Manejo de Caja / Vidriera / Adelanto
    manejo_caja: bool = False,
    cajero_tipo: str = "",
    faltante_caja: float = 0,
    armado_vidriera: bool = False,
    adelanto_sueldo: float = 0,
    adelanto_vacaciones: float = 0,
    sac_prop_mes: bool = False,
    sac_base_rem: float = -1,
    sac_base_nr: float = -1,
    sac_factor: float = 1,
    sac_base_period: str = "",
    # Agua potable: selector A/B/C/D. Se mantiene conexiones por compatibilidad.
    conex_cat: str = "",
    conexiones: int = 0,
    # Fúnebres: ids de adicionales seleccionados (coma-separados)
    fun_adic: Optional[List[str]] = Query(None),
    # Ley 27.802 / art. 140 LCT: conceptos a cargo del empleador (desde 2026-05)
    regimen_contribuciones: str = "inciso_b",
    art_pct: float = 3,
    art_fijo: float = 1765,
    scvo_legal: bool = True,
    seguro_vida_cct_prima: float = 0,
    osecac_adicional_patronal: bool = True,
    la_estrella: bool = True,
    instituto_capacitacion: bool = True,
):
    return calcular_payload(
        rama=rama,
        agrup=agrup,
        categoria=categoria,
        mes=mes,
        jornada=jornada,
        basico_manual=basico_manual,
        anios_antig=anios_antig,
        osecac=osecac,
        obra_social_sobre_no_rem=obra_social_sobre_no_rem,
        afiliado=afiliado,
        sind_pct=sind_pct,
        sind_fijo=sind_fijo,
        titulo_pct=titulo_pct,
        zona_pct=zona_pct,
        fer_no_trab=fer_no_trab,
        fer_trab=fer_trab,
        vac_goz=vac_goz,
        aus_inj=aus_inj,
        jubilado=jubilado,
        susp_dias=susp_dias,
        embargo=embargo,
        hex50=hex50,
        hex100=hex100,
        hs_noct=hs_noct,
        km_tipo=km_tipo,
        km_menos100=km_menos100,
        km_mas100=km_mas100,
        a_cuenta_rem=a_cuenta_rem,
        viaticos_nr=viaticos_nr,
        manejo_caja=manejo_caja,
        cajero_tipo=cajero_tipo,
        faltante_caja=faltante_caja,
        armado_vidriera=armado_vidriera,
        adelanto_sueldo=adelanto_sueldo,
        adelanto_vacaciones=adelanto_vacaciones,
        sac_prop_mes=sac_prop_mes,
        sac_base_rem=sac_base_rem,
        sac_base_nr=sac_base_nr,
        sac_factor=sac_factor,
        sac_base_period=sac_base_period,
        conex_cat=conex_cat,
        conexiones=conexiones,
        fun_adic=(";".join(fun_adic) if fun_adic else ""),
        regimen_contribuciones=regimen_contribuciones,
        art_pct=art_pct,
        art_fijo=art_fijo,
        scvo_legal=scvo_legal,
        seguro_vida_cct_prima=seguro_vida_cct_prima,
        osecac_adicional_patronal=osecac_adicional_patronal,
        la_estrella=la_estrella,
        instituto_capacitacion=instituto_capacitacion,
    )



# ========= VACACIONES EMPRESAS =========
@app.get("/calcular-vacaciones")
def calcular_vacaciones(
    rama: str,
    agrup: str,
    categoria: str,
    mes: str,
    dias: float,
    base_rem: float,
    base_nr: float = 0,
    osecac: bool = True,
    afiliado: bool = False,
    sind_pct: float = 0,
    jubilado: bool = False,
    regimen_contribuciones: str = "inciso_b",
    art_pct: float = 3,
    art_fijo: float = 1765,
):
    return calcular_vacaciones_payload(
        rama=rama,
        agrup=agrup,
        categoria=categoria,
        mes=mes,
        dias=dias,
        base_rem=base_rem,
        base_nr=base_nr,
        osecac=osecac,
        afiliado=afiliado,
        sind_pct=sind_pct,
        jubilado=jubilado,
        regimen_contribuciones=regimen_contribuciones,
        art_pct=art_pct,
        art_fijo=art_fijo,
    )


# ========= CALCULAR FINAL (liquidación final) =========
@app.get("/calcular-final")
def calcular_final(
    rama: str,
    agrup: str,
    categoria: str,
    fecha_ingreso: str,
    fecha_egreso: str,
    jornada: float = 48.0,
    tipo: str = "RENUNCIA",
    # Mejor salario mensual normal y habitual (ideal: desglosado)
    mejor_rem: float = 0,
    mejor_nr: float = 0,
    mejor_total: float = 0,
    # Parámetros
    dias_mes: int = 0,
    vac_anuales: int = 14,
    vac_dias_computables: float = 0.0,
    vac_no_gozadas_dias: float = 0.0,
    preaviso_dias: int = 0,
    integracion: bool = True,
    sac_preaviso: bool = False,
    sac_integracion: bool = True,
    sac_devengado_rem: float = -1,
    sac_devengado_nr: float = -1,
    # Mismos flags/descuentos que mensual
    osecac: bool = True,
    afiliado: bool = False,
    sind_pct: float = 0,
    sind_fijo: float = 0,
    titulo_pct: float = 0,
    zona_pct: float = 0,
    fer_no_trab: int = 0,
    fer_trab: int = 0,
    vac_goz: int = 0,
    aus_inj: int = 0,
    susp_dias: int = 0,
    hex50: float = 0,
    hex100: float = 0,
    hs_noct: float = 0,
    km_tipo: str = "",
    km_menos100: int = 0,
    km_mas100: int = 0,
    a_cuenta_rem: float = 0,
    viaticos_nr: float = 0,
    manejo_caja: bool = False,
    cajero_tipo: str = "",
    faltante_caja: float = 0,
    armado_vidriera: bool = False,
    adelanto_sueldo: float = 0,
    fun_adic: Optional[List[str]] = Query(default=[]),
    jubilado: bool = False,
    embargo: float = 0,
    regimen_contribuciones: str = "inciso_b",
    art_pct: float = 3,
    art_fijo: float = 1765,
    scvo_legal: bool = True,
    seguro_vida_cct_prima: float = 0,
    osecac_adicional_patronal: bool = True,
    la_estrella: bool = True,
    instituto_capacitacion: bool = True,
    authorization: Optional[str] = Header(default=None),
):
    return calcular_final_payload(
        rama=rama,
        agrup=agrup,
        categoria=categoria,
        jornada=jornada,
        fecha_ingreso=fecha_ingreso,
        fecha_egreso=fecha_egreso,
        tipo=tipo,
        mejor_rem=mejor_rem,
        mejor_nr=mejor_nr,
        mejor_total=mejor_total,
        dias_mes=dias_mes,
        vac_anuales=vac_anuales,
        vac_dias_computables=vac_dias_computables,
        vac_no_gozadas_dias=vac_no_gozadas_dias,
        preaviso_dias=preaviso_dias,
        integracion=integracion,
        sac_sobre_preaviso=sac_preaviso,
        sac_sobre_integracion=sac_integracion,
        sac_devengado_rem=sac_devengado_rem,
        sac_devengado_nr=sac_devengado_nr,
        osecac=osecac,
        afiliado=afiliado,
        sind_pct=sind_pct,
        sind_fijo=sind_fijo,
        titulo_pct=titulo_pct,
        zona_pct=zona_pct,
        fer_no_trab=fer_no_trab,
        fer_trab=fer_trab,
        vac_goz=vac_goz,
        aus_inj=aus_inj,
        susp_dias=susp_dias,
        hex50=hex50,
        hex100=hex100,
        hs_noct=hs_noct,
        km_tipo=km_tipo,
        km_menos100=km_menos100,
        km_mas100=km_mas100,
        a_cuenta_rem=a_cuenta_rem,
        viaticos_nr=viaticos_nr,
        manejo_caja=manejo_caja,
        cajero_tipo=cajero_tipo,
        faltante_caja=faltante_caja,
        armado_vidriera=armado_vidriera,
        adelanto_sueldo=adelanto_sueldo,
        fun_adic=(";".join(fun_adic) if fun_adic else ""),
        jubilado=jubilado,
        embargo=embargo,
        regimen_contribuciones=regimen_contribuciones,
        art_pct=art_pct,
        art_fijo=art_fijo,
        scvo_legal=scvo_legal,
        seguro_vida_cct_prima=seguro_vida_cct_prima,
        osecac_adicional_patronal=osecac_adicional_patronal,
        la_estrella=la_estrella,
        instituto_capacitacion=instituto_capacitacion,
    )
# ========= FUNEBRES =========
@app.get("/adicionales-funebres")
def adicionales_funebres(mes: str):
    return get_adicionales_funebres(mes)

# ========= AGUA POTABLE =========
@app.get("/regla-conexiones")
def regla_conexiones(cantidad: int = 0, nivel: str = ""):
    # Si el front manda nivel (A/B/C/D), devolvemos la misma estructura.
    if nivel:
        return match_regla_conexiones(nivel)
    return match_regla_conexiones(cantidad)

# ========= TURISMO =========
@app.get("/titulo-pct")
def titulo_pct(nivel: str):
    return get_titulo_pct_por_nivel(nivel)

# ========= CAJEROS =========
@app.get("/regla-cajero")
def regla_cajero(tipo: str):
    return get_regla_cajero(tipo)

# ========= KM =========
@app.get("/regla-km")
def regla_km(categoria: str, km: float):
    return get_regla_km(categoria, km)

