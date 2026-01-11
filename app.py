
from __future__ import annotations

import os
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, Query
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
MAESTRO_PATH = os.path.join(APP_ROOT, "data", "maestro_actualizado.xlsx")

app = FastAPI(title="Motor Sueldos FAECYS")

# Servir el frontend
if os.path.isdir(os.path.join(APP_ROOT, "public")):
    app.mount("/", StaticFiles(directory=os.path.join(APP_ROOT, "public"), html=True), name="public")


def _norm(s: Any) -> str:
    return " ".join(str(s or "").strip().split())


def _month_key(m: str) -> str:
    # Mantener formato YYYY-MM si viene así; si viene fecha YYYY-MM-DD, cortar
    m = _norm(m)
    if len(m) >= 7 and m[4] == "-" and m[6].isdigit():
        return m[:7]
    return m


@lru_cache(maxsize=1)
def _load_rows() -> List[Dict[str, Any]]:
    if not os.path.exists(MAESTRO_PATH):
        raise FileNotFoundError(f"No se encontró el maestro en {MAESTRO_PATH}")

    wb = load_workbook(MAESTRO_PATH, data_only=True)

    rows: List[Dict[str, Any]] = []
    for name in wb.sheetnames:
        if not name.startswith("Categorias_"):
            continue
        ws = wb[name]
        # Esperamos encabezados: Rama, Agrupamiento, Categoria, Mes, Basico, No Remunerativo, SUMA_FIJA
        header = [(_norm(c) or "").lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(header) if h}
        required = ["rama", "agrupamiento", "categoria", "mes", "basico", "no remunerativo", "suma_fija"]
        # permitir variantes de suma fija
        if "suma_fija" not in idx and "suma fija" in idx:
            idx["suma_fija"] = idx["suma fija"]
        if "no remunerativo" not in idx and "no remunerativo" in idx:
            idx["no remunerativo"] = idx["no remunerativo"]

        missing = [k for k in ["rama", "agrupamiento", "categoria", "mes", "basico"] if k not in idx]
        if missing:
            # hoja distinta; la ignoramos para no romper
            continue

        for r in ws.iter_rows(min_row=2, values_only=True):
            rama = _norm(r[idx["rama"]])
            agrup = _norm(r[idx["agrupamiento"]])
            cat = _norm(r[idx["categoria"]])
            mes = _month_key(r[idx["mes"]])
            basico = float(r[idx["basico"]] or 0)
            no_rem = float(r[idx.get("no remunerativo", -1)] or 0) if idx.get("no remunerativo", -1) != -1 else 0.0
            suma_fija = float(r[idx.get("suma_fija", -1)] or 0) if idx.get("suma_fija", -1) != -1 else 0.0

            if not rama or not mes or not cat:
                continue

            rows.append({
                "rama": rama,
                "agrup": agrup or "—",
                "categoria": cat,
                "mes": mes,
                "basico": basico,
                "no_rem_1": no_rem,
                "no_rem_2": suma_fija,
            })

    return rows


def _index_rows(rows: List[Dict[str, Any]]) -> Dict[Tuple[str, str, str, str], Dict[str, Any]]:
    out: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}
    for it in rows:
        key = (_norm(it["rama"]).upper(), _norm(it["agrup"]), _norm(it["categoria"]), _month_key(it["mes"]))
        out[key] = it
    return out


@lru_cache(maxsize=1)
def _rows_index():
    rows = _load_rows()
    return _index_rows(rows)


@app.get("/health")
def health():
    return {"ok": True}


@app.get("/meta")
def meta():
    try:
        rows = _load_rows()
        ramas = sorted({r["rama"] for r in rows})
        meses = sorted({r["mes"] for r in rows})
        return {"ok": True, "ramas": ramas, "meses": meses}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/payload")
def payload():
    # El frontend lo usa para armar el árbol (buildTree)
    try:
        rows = _load_rows()
        # devolver solo lo necesario + valores por si se usan
        return {"ok": True, "rows": rows}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/calcular")
def calcular(
    rama: str = Query(...),
    mes: str = Query(...),
    agrup: str = Query(...),
    categoria: str = Query(...),
):
    try:
        key = (_norm(rama).upper(), _norm(agrup), _norm(categoria), _month_key(mes))
        hit = _rows_index().get(key)
        if not hit:
            # fallback: intentar con agrup "—" si viene vacío
            key2 = (_norm(rama).upper(), "—", _norm(categoria), _month_key(mes))
            hit = _rows_index().get(key2)
        if not hit:
            return {"ok": False, "error": "No se encontró esa combinación en el maestro", "rama": rama, "agrup": agrup, "categoria": categoria, "mes": mes}

        return {
            "ok": True,
            "basico": hit["basico"],
            "no_rem_1": hit["no_rem_1"],
            "no_rem_2": hit["no_rem_2"],
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})
