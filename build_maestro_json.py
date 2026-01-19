import json
import re
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parents[1] / "maestro_actualizado.xlsx"
OUT = Path(__file__).resolve().parent / "data" / "maestro.json"


def norm(h):
    if h is None:
        return None
    h = re.sub(r"\s+", " ", str(h).strip())
    return h


def export_sheet_rows(ws, required):
    headers = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    h2i = {h: i + 1 for i, h in enumerate(headers) if h}
    required_norm = {}
    for want in required:
        for h in headers:
            if not h:
                continue
            if h.strip().lower() == want.strip().lower():
                required_norm[want] = h
                break
    missing = [w for w in required if w not in required_norm]
    if missing:
        return []

    out = []
    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for want, real in required_norm.items():
            v = ws.cell(r, h2i[real]).value
            if v not in (None, ""):
                empty = False
            row[want] = v
        if not empty:
            out.append(row)
    return out


def main():
    if not SRC.exists():
        raise SystemExit(f"No existe {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True)

    escala = []
    for name in wb.sheetnames:
        if name.startswith("Categorias_"):
            ws = wb[name]
            rows = export_sheet_rows(ws, ["Rama", "Agrupamiento", "Categoria", "Mes", "Basico", "No Remunerativo", "SUMA_FIJA"])
            escala.extend(rows)

    adicionales = []
    if "Adicionales" in wb.sheetnames:
        adicionales = export_sheet_rows(wb["Adicionales"], ["Rama", "Concepto", "Mes", "Valor", "Detalle"])

    reglas = []
    if "ReglasAdicionales" in wb.sheetnames:
        reglas = export_sheet_rows(
            wb["ReglasAdicionales"],
            ["regla_id", "rama_aplica", "concepto", "articulo", "tipo", "parametro", "tramo_desde", "tramo_hasta", "porcentaje",
             "base_ref_rama", "base_ref_agrup", "base_ref_categoria", "base_ref_mes", "observaciones"],
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"escala": escala, "adicionales": adicionales, "reglas_adicionales": reglas}, ensure_ascii=False), encoding="utf-8")
    print(f"OK -> {OUT} (escala={len(escala)})")


if __name__ == "__main__":
    main()
